from __future__ import annotations

import asyncio
import csv
import logging
import os
import tempfile
from datetime import datetime, date as dt_date
from typing import Any, AsyncIterator, Tuple

import yadisk
from openpyxl import load_workbook

from repositories.yandex import get_yandex_repo

logger = logging.getLogger(__name__)


def _is_dir(item: Any) -> bool:
    # В yadisk обычно item.type == "dir"
    t = getattr(item, "type", None)
    if t is not None:
        return t == "dir"
    # На всякий случай — если есть метод/флаг
    is_dir = getattr(item, "is_dir", None)
    return bool(is_dir() if callable(is_dir) else is_dir)


def _item_path(parent: str, item: Any) -> str:
    # Обычно есть item.path (часто вида "disk:/..."), и его можно передавать обратно в методы.
    p = getattr(item, "path", None)
    if p:
        return p
    name = getattr(item, "name", None) or str(item)
    return parent.rstrip("/") + "/" + name


async def walk_all(
    client: yadisk.AsyncClient,
    root: str,
) -> AsyncIterator[Tuple[str, Any]]:
    """
    Yield: (full_path, item_resource)
    """
    stack = [root]
    while stack:
        cur = stack.pop()
        async for item in client.listdir(cur):
            full = _item_path(cur, item)
            yield full, item
            if _is_dir(item):
                stack.append(full)


repo = get_yandex_repo()

report_path = os.getenv("REPORT_PATH")
TOKEN = os.getenv("YADISK_TOKEN")
start_date = datetime.strptime("2020-01-01", "%Y-%m-%d")
end_date = datetime.now()


async def main():
    tariffs = set()
    dates = []
    result = []

    tariff_config = {}
    skipped_tariffs = []
    with open("tariff_config.csv") as csvfile:
        reader = csv.DictReader(csvfile, delimiter=";")
        for row in reader:
            for group, tariff_name in row.items():
                group = group.strip()
                tariff_name = tariff_name.strip()
                if tariff_name:
                    group_data = tariff_config.setdefault(group, [])
                    group_data.append(tariff_name)
                    skipped_tariffs.append(tariff_name)

    async with yadisk.AsyncClient(token=TOKEN) as client:
        ok = await client.check_token()
        if not ok:
            pass

        async for path, item in walk_all(client, report_path):
            if getattr(item, "type", None) == "file" and path.endswith(".xlsx") and "Итоговый отчет" in item.name:
                try:
                    report_date = datetime.strptime(item.name.split(" ")[0], "%Y-%m-%d")
                except ValueError:
                    continue

                if report_date < start_date:
                    continue

                if report_date > end_date:
                    continue

                with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
                    for _ in range(10):
                        try:
                            await client.download(path, tmp.name)
                        except yadisk.exceptions.RequestError:
                            logger.error("Request error. Try again later.")
                            await asyncio.sleep(3)
                            continue
                        else:
                            break

                    wb = load_workbook(tmp.name, data_only=True)
                    ws = wb.active
                    result_line = {}
                    for row in ws.iter_rows(min_row=1):
                        tariff_field = row[1].value if len(row) > 1 else None
                        date_field = row[4].value if len(row) > 4 else None
                        count_field = row[9].value if len(row) > 9 else None
                        sum_field = row[11].value if len(row) > 11 else None
                        if tariff_field and isinstance(count_field, int) and isinstance(sum_field, int):
                            tariffs.add(tariff_field)

                            for group, tariff_collection in tariff_config.items():
                                if tariff_field in tariff_collection:
                                    result_line.setdefault(group, 0)
                                    result_line[group] += count_field
                                    try:
                                        skipped_tariffs.remove(tariff_field)
                                    except ValueError:
                                        pass

                        if date_field:
                            dates.append(date_field)

                            result_line["Дата"] = date_field

                    if len(result_line) < 2:
                        continue

                    result.append(result_line)

        result = sorted(result, key=lambda x: datetime.strptime(x.get("Дата"), "%d.%m.%Y"))
        for line in result:
            for group in tariff_config:
                if group not in line:
                    line[group] = 0

        with open("result.csv", "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=result[0].keys())
            writer.writeheader()
            writer.writerows(result)

        remote_csv_path = f"{report_path.rstrip('/')}/result_from_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.csv"
        await client.upload("result.csv", remote_csv_path, overwrite=True)
        logger.info("CSV загружен на Яндекс.Диск: %s", remote_csv_path)

        file_name = f"skipped_tariffs_from_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.txt"
        remote_path = f"{report_path.rstrip('/')}/{file_name}"
        with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", encoding="utf-8") as tmp_txt:
            tmp_txt.write("\n".join(skipped_tariffs))
            tmp_txt.flush()
            await client.upload(tmp_txt.name, remote_path, overwrite=True)

        # tariffs_list = sorted(map(str, tariffs))
        # file_name = f"tariffs from {start_date.strftime('%Y-%m-%d')} ({len(tariffs_list)}).txt"
        # remote_path = f"{report_path.rstrip('/')}/{file_name}"
        # with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", encoding="utf-8") as tmp_txt:
        #     tmp_txt.write("\n".join(tariffs_list))
        #     tmp_txt.flush()
        #     await client.upload(tmp_txt.name, remote_path, overwrite=True)
        #
        # sorted_dates = sorted(dates)
        # file_name = f"dates from {start_date.strftime('%Y-%m-%d')}.txt"
        # remote_path = f"{report_path.rstrip('/')}/{file_name}"
        # with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", encoding="utf-8") as tmp_txt:
        #     tmp_txt.write("\n".join(sorted_dates))
        #     tmp_txt.flush()
        #     await client.upload(tmp_txt.name, remote_path, overwrite=True)
        #
        # exclude_dates = []
        #
        # def get_next_date(_start_date: datetime) -> GeneratorExit[str]:
        #     current_date = _start_date
        #     while True:
        #         yield datetime.strftime(current_date, "%Y-%m-%d")
        #         current_date += timedelta(days=1)
        #         if current_date > datetime.now():
        #             break
        #
        # for date in get_next_date(start_date):
        #     if date not in dates:
        #         exclude_dates.append(date)
        #
        # sorted_exclude_dates = sorted(exclude_dates)
        # file_name = f"exclude_dates from {start_date.strftime('%Y-%m-%d')}.txt"
        # remote_path = f"{report_path.rstrip('/')}/{file_name}"
        # with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", encoding="utf-8") as tmp_txt:
        #     tmp_txt.write("\n".join(sorted_exclude_dates))
        #     tmp_txt.flush()
        #     await client.upload(tmp_txt.name, remote_path, overwrite=True)


asyncio.run(main())
