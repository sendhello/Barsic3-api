import logging
import re
from datetime import datetime, timedelta
from decimal import Decimal

import apiclient
import httplib2
from dateutil.relativedelta import relativedelta
from fastapi.exceptions import HTTPException
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from starlette import status

from core.settings import settings
from db.mssql import MsSqlDatabase
from legacy import functions
from legacy.to_google_sheets import Spreadsheet, create_new_google_doc
from repositories.yandex import YandexRepository, get_yandex_repo
from schemas.bars import ClientsCount
from schemas.google_report_ids import GoogleReportIdCreate
from services.bars import BarsService, get_bars_service
from services.report_config import ReportConfigService, get_report_config_service
from services.rk import RKService, get_rk_service
from services.settings import SettingsService, get_settings_service
from sql.clients_count import CLIENTS_COUNT_SQL


logger = logging.getLogger("barsicreport2")


class BarsicReport2Service:
    """
    Функционал предыдущей версии.
    """

    def __init__(self, bars_srv: MsSqlDatabase, rk_srv: MsSqlDatabase):
        self.bars_srv = bars_srv
        self.rk_srv = rk_srv

        self.org1 = None
        self.org2 = None
        self.org3 = None
        self.org4 = None
        self.org5 = None

        self.count_sql_error = 0
        self.org_for_finreport = {}
        self.orgs = []
        self.new_agentservice = []
        self.agentorgs = []

        self._report_config_service: ReportConfigService = get_report_config_service()
        self._settings_service: SettingsService = get_settings_service()
        self._bars_service: BarsService = get_bars_service()
        self._rk_service: RKService = get_rk_service()
        self._yandex_repo: YandexRepository = get_yandex_repo()

    def get_clients_count(self) -> list[ClientsCount]:
        """Получение количества человек в зоне."""

        self.bars_srv.set_database(settings.mssql_database1)
        with self.bars_srv as connect:
            cursor = connect.cursor()
            cursor.execute(CLIENTS_COUNT_SQL)
            rows = cursor.fetchall()
            if not rows:
                return [ClientsCount(count=0, id=488, zone_name="", code="0003")]

        return [
            ClientsCount(count=row[0], id=row[1], zone_name=row[2], code=row[3])
            for row in rows
        ]

    def count_clients_print(self):
        clients_count = self.get_clients_count()
        self.click_select_org()
        self.bars_srv.set_database(settings.mssql_database1)
        with self.bars_srv as connect:
            total_report = functions.get_total_report(
                connect=connect,
                org=self.org1[0],
                org_name=self.org1[1],
                date_from=datetime.now(),
                date_to=datetime.now() + timedelta(1),
            )
        try:
            count_clients = int(total_report["Аквазона"][0])
        except KeyError:
            count_clients = 0

        try:
            count_clients_allday = self.reportClientCountTotals(
                database=settings.mssql_database1,
                org=self.org1[0],
                date_from=datetime.now(),
                date_to=datetime.now() + timedelta(1),
            )[0][1]
        except IndexError:
            count_clients_allday = 0

        return {
            "Всего": str(count_clients) + " / " + str(count_clients_allday),
            clients_count[0].zone_name: clients_count[0].count,
        }

    def click_select_org(self):
        """
        Выбор первой организации из списка организаций
        """
        org_list1 = self.list_organisation(
            database=settings.mssql_database1,
        )
        org_list2 = self.list_organisation(
            database=settings.mssql_database2,
        )
        for org in org_list1:
            if org[0] == 36:
                self.org1 = (org[0], org[2])
            if org[0] == 7203673:
                self.org3 = (org[0], org[2])
            if org[0] == 7203674:
                self.org4 = (org[0], org[2])
            if org[0] == 13240081:
                self.org5 = (org[0], org[2])

        self.org2 = (org_list2[0][0], org_list2[0][2])
        logger.info(f"Выбраны организации {org_list1[0][2]} и {org_list2[0][2]}")

    def list_organisation(
        self,
        database,
    ):
        """Функция делает запрос в базу Барс и возвращает список заведенных
        в базе организаций в виде списка кортежей."""

        self.bars_srv.set_database(database)
        with self.bars_srv as connect:
            cursor = connect.cursor()
            id_type = 1
            cursor.execute(
                f"""
                SELECT
                    SuperAccountId, Type, Descr, CanRegister, CanPass, IsStuff, IsBlocked, 
                    BlockReason, DenyReturn, ClientCategoryId, DiscountCard, PersonalInfoId, 
                    Address, Inn, ExternalId, RegisterTime,LastTransactionTime, 
                    LegalEntityRelationTypeId, SellServicePointId, DepositServicePointId, 
                    AllowIgnoreStoredPledge, Email, Latitude, Longitude, Phone, WebSite, 
                    TNG_ProfileId
                FROM
                    SuperAccount
                WHERE
                    Type={id_type}
                """
            )
            rows = cursor.fetchall()

        return rows

    def reportClientCountTotals(
        self,
        database,
        org,
        date_from,
        date_to,
    ):
        date_from = date_from.strftime("%Y%m%d 00:00:00")
        date_to = date_to.strftime("%Y%m%d 00:00:00")

        self.bars_srv.set_database(database)
        with self.bars_srv as connect:
            cursor = connect.cursor()
            cursor.execute(
                f"exec sp_reportClientCountTotals @sa={org},@from='{date_from}',@to='{date_to}',@categoryId=0"
            )
            rows = cursor.fetchall()

        return rows

    def client_count_totals_period(
        self,
        database,
        org,
        org_name,
        date_from,
        date_to,
    ):
        """
        Если выбран 1 день возвращает словарь количества людей за текущий месяц,
        если выбран период возвращает словарь количества людей за период
        """
        count = []

        if date_from + timedelta(1) == date_to:
            first_day = datetime.strptime((date_from.strftime("%Y%m") + "01"), "%Y%m%d")
            count.append((org_name, 1))
        else:
            first_day = date_from
            count.append((org_name, 0))
        total = 0
        while first_day < date_to:
            client_count = self.reportClientCountTotals(
                database=database,
                org=org,
                date_from=first_day,
                date_to=first_day + timedelta(1),
            )
            try:
                count.append((client_count[0][0], client_count[0][1]))
                total += client_count[0][1]
            except IndexError:
                count.append((first_day, 0))
            first_day += timedelta(1)
        count.append(("Итого", total))
        return count

    def cash_report_request(
        self,
        database,
        date_from,
        date_to,
    ):
        """Делает запрос в базу Барс и возвращает суммовой отчет за запрашиваемый период."""

        date_from = date_from.strftime("%Y%m%d 00:00:00")
        date_to = date_to.strftime("%Y%m%d 00:00:00")

        self.bars_srv.set_database(database)
        with self.bars_srv as connect:
            cursor = connect.cursor()
            cursor.execute(
                f"exec sp_reportCashDeskMoney @from='{date_from}', @to='{date_to}'"
            )
            rows = cursor.fetchall()

        return rows

    def service_point_request(
        self,
        database,
    ):
        """Делает запрос в базу Барс и возвращает список рабочих мест."""

        self.bars_srv.set_database(database)
        with self.bars_srv as connect:
            cursor = connect.cursor()
            cursor.execute(
                """
                    SELECT
                        ServicePointId, Name, SuperAccountId, Type, Code, IsInternal
                    FROM 
                        ServicePoint
                """
            )
            rows = cursor.fetchall()

        return rows

    def cashdesk_report(
        self,
        database,
        date_from,
        date_to,
    ):
        """
        Преобразует запросы из базы в суммовой отчет
        :return: dict
        """
        cash_report = self.cash_report_request(
            database=database,
            date_from=date_from,
            date_to=date_to,
        )
        service_point = self.service_point_request(
            database=database,
        )
        service_point_dict = {}
        for point in service_point:
            service_point_dict[point[0]] = (
                point[1],
                point[2],
                point[3],
                point[4],
                point[5],
            )
        report = {}
        for line in cash_report:
            report[line[8]] = []
        for line in cash_report:
            report[line[8]].append(
                [
                    service_point_dict[line[0]][0],
                    line[1],
                    line[2],
                    line[3],
                    line[4],
                    line[5],
                    line[6],
                    line[7],
                ]
            )
        all_sum = [
            "Итого по отчету",
            Decimal(0.0),
            Decimal(0.0),
            Decimal(0.0),
            Decimal(0.0),
            Decimal(0.0),
            Decimal(0.0),
            Decimal(0.0),
        ]
        for typpe in report:
            type_sum = [
                "Итого",
                Decimal(0.0),
                Decimal(0.0),
                Decimal(0.0),
                Decimal(0.0),
                Decimal(0.0),
                Decimal(0.0),
                Decimal(0.0),
            ]
            for line in report[typpe]:
                i = 0
                while True:
                    i += 1
                    try:
                        type_sum[i] += line[i]
                        all_sum[i] += line[i]
                    except IndexError:
                        break
            report[typpe].append(type_sum)
        report["Итого"] = [all_sum]
        report["Дата"] = [[date_from, date_to]]
        if database == settings.mssql_database1:
            report["Организация"] = [[self.org1[1]]]
        elif database == settings.mssql_database2:
            report["Организация"] = [[self.org2[1]]]
        return report

    def fin_report(self):
        """
        Форминует финансовый отчет в установленном формате
        :return - dict
        """
        logger.info("Формирование финансового отчета")
        self.finreport_dict = {}
        is_aquazona = None
        for org, services in self.orgs_dict.items():
            if org != "Не учитывать":
                self.finreport_dict[org] = [0, 0.00]
                for serv in services:
                    try:
                        if org == "Дата":
                            self.finreport_dict[org][0] = self.itog_report_org1[serv][0]
                            self.finreport_dict[org][1] = self.itog_report_org1[serv][1]
                        elif serv == "Депозит":
                            self.finreport_dict[org][1] += self.itog_report_org1[serv][
                                1
                            ]
                        elif serv == "Аквазона":
                            self.finreport_dict["Кол-во проходов"] = [
                                self.itog_report_org1[serv][0],
                                0,
                            ]
                            self.finreport_dict[org][1] += self.itog_report_org1[serv][
                                1
                            ]
                            is_aquazona = True
                        elif serv == "Организация":
                            pass
                        else:
                            if (
                                self.itog_report_org1.get(serv)
                                and self.itog_report_org1[serv][1] != 0.0
                            ):
                                self.finreport_dict[org][0] += self.itog_report_org1[
                                    serv
                                ][0]
                                self.finreport_dict[org][1] += self.itog_report_org1[
                                    serv
                                ][1]
                            if (
                                self.itog_report_org3.get(serv)
                                and self.itog_report_org3[serv][1] != 0.0
                            ):
                                self.finreport_dict[org][0] += self.itog_report_org3[
                                    serv
                                ][0]
                                self.finreport_dict[org][1] += self.itog_report_org3[
                                    serv
                                ][1]
                            if (
                                self.itog_report_org4.get(serv)
                                and self.itog_report_org4[serv][1] != 0.0
                            ):
                                self.finreport_dict[org][0] += self.itog_report_org4[
                                    serv
                                ][0]
                                self.finreport_dict[org][1] += self.itog_report_org4[
                                    serv
                                ][1]
                            if (
                                self.itog_report_org5.get(serv)
                                and self.itog_report_org5[serv][1] != 0.0
                            ):
                                self.finreport_dict[org][0] += self.itog_report_org5[
                                    serv
                                ][0]
                                self.finreport_dict[org][1] += self.itog_report_org5[
                                    serv
                                ][1]
                    except KeyError:
                        pass
                    except TypeError:
                        pass

        if not is_aquazona:
            self.finreport_dict["Кол-во проходов"] = [0, 0.00]

        self.finreport_dict.setdefault("Online Продажи", [0, 0.0])
        self.finreport_dict["Online Продажи"][0] += self.report_bitrix[0]
        self.finreport_dict["Online Продажи"][1] += self.report_bitrix[1]

        self.finreport_dict["Смайл"][0] = self.smile_report.total_count
        self.finreport_dict["Смайл"][1] = self.smile_report.total_sum

        total_cashdesk_report = self.cashdesk_report_org1["Итого"][0]
        self.finreport_dict["MaxBonus"] = (
            0,
            float(total_cashdesk_report[6] - total_cashdesk_report[7]),
        )

    def fin_report_lastyear(self):
        """
        Форминует финансовый отчет за предыдущий год в установленном формате
        :return - dict
        """
        logger.info("Формирование финансового отчета за прошлый год")
        self.finreport_dict_lastyear = {}
        is_aquazona = None
        for org, services in self.orgs_dict.items():
            if org != "Не учитывать":
                self.finreport_dict_lastyear[org] = [0, 0.00]
                for serv in services:
                    try:
                        if org == "Дата":
                            self.finreport_dict_lastyear[org][0] = (
                                self.itog_report_org1_lastyear[serv][0]
                            )
                            self.finreport_dict_lastyear[org][1] = (
                                self.itog_report_org1_lastyear[serv][1]
                            )
                        elif serv == "Депозит":
                            self.finreport_dict_lastyear[org][
                                1
                            ] += self.itog_report_org1_lastyear[serv][1]
                        elif serv == "Аквазона":
                            self.finreport_dict_lastyear["Кол-во проходов"] = [
                                self.itog_report_org1_lastyear[serv][0],
                                0,
                            ]
                            self.finreport_dict_lastyear[org][
                                1
                            ] += self.itog_report_org1_lastyear[serv][1]
                            is_aquazona = True
                        elif serv == "Организация":
                            pass
                        else:
                            if (
                                self.itog_report_org1_lastyear.get(serv)
                                and self.itog_report_org1_lastyear[serv][1] != 0.0
                            ):
                                self.finreport_dict_lastyear[org][
                                    0
                                ] += self.itog_report_org1_lastyear[serv][0]
                                self.finreport_dict_lastyear[org][
                                    1
                                ] += self.itog_report_org1_lastyear[serv][1]
                            if (
                                self.itog_report_org3_lastyear.get(serv)
                                and self.itog_report_org3_lastyear[serv][1] != 0.0
                            ):
                                self.finreport_dict_lastyear[org][
                                    0
                                ] += self.itog_report_org3_lastyear[serv][0]
                                self.finreport_dict_lastyear[org][
                                    1
                                ] += self.itog_report_org3_lastyear[serv][1]
                            if (
                                self.itog_report_org4_lastyear.get(serv)
                                and self.itog_report_org4_lastyear[serv][1] != 0.0
                            ):
                                self.finreport_dict_lastyear[org][
                                    0
                                ] += self.itog_report_org4_lastyear[serv][0]
                                self.finreport_dict_lastyear[org][
                                    1
                                ] += self.itog_report_org4_lastyear[serv][1]
                            if (
                                self.itog_report_org5_lastyear.get(serv)
                                and self.itog_report_org5_lastyear[serv][1] != 0.0
                            ):
                                self.finreport_dict_lastyear[org][
                                    0
                                ] += self.itog_report_org5_lastyear[serv][0]
                                self.finreport_dict_lastyear[org][
                                    1
                                ] += self.itog_report_org5_lastyear[serv][1]
                    except KeyError:
                        pass
                    except TypeError:
                        pass
        if not is_aquazona:
            self.finreport_dict_lastyear["Кол-во проходов"] = [0, 0.00]
        self.finreport_dict_lastyear.setdefault("Online Продажи", [0, 0.0])
        self.finreport_dict_lastyear["Online Продажи"][
            0
        ] += self.report_bitrix_lastyear[0]
        self.finreport_dict_lastyear["Online Продажи"][
            1
        ] += self.report_bitrix_lastyear[1]
        self.finreport_dict_lastyear["Смайл"][
            0
        ] = self.smile_report_lastyear.total_count
        self.finreport_dict_lastyear["Смайл"][1] = self.smile_report_lastyear.total_sum

        total_cashdesk_report = self.cashdesk_report_org1_lastyear["Итого"][0]
        self.finreport_dict_lastyear["MaxBonus"] = (
            0,
            total_cashdesk_report[6] - total_cashdesk_report[7],
        )

    def fin_report_beach(self):
        """
        Форминует финансовый отчет по пляжу в установленном формате
        :return - dict
        """
        logger.info("Формирование финансового отчета по пляжу")
        self.finreport_dict_beach = {
            "Депозит": (0, 0),
            "Товары": (0, 0),
            "Услуги": (0, 0),
            "Карты": (0, 0),
            "Итого по отчету": (0, 0),
        }
        for service in self.itog_report_org2:
            if service == "Дата":
                self.finreport_dict_beach[service] = (
                    self.itog_report_org2[service][0],
                    self.itog_report_org2[service][1],
                )
            elif service == "Выход с пляжа":
                self.finreport_dict_beach[service] = (
                    self.itog_report_org2[service][0],
                    self.itog_report_org2[service][1],
                )
            elif not self.itog_report_org2[service][3] in self.finreport_dict_beach:
                self.finreport_dict_beach[self.itog_report_org2[service][3]] = (
                    self.itog_report_org2[service][0],
                    self.itog_report_org2[service][1],
                )
            else:
                try:
                    self.finreport_dict_beach[self.itog_report_org2[service][3]] = (
                        self.finreport_dict_beach[self.itog_report_org2[service][3]][0]
                        + self.itog_report_org2[service][0],
                        self.finreport_dict_beach[self.itog_report_org2[service][3]][1]
                        + self.itog_report_org2[service][1],
                    )
                except TypeError:
                    pass
        if "Выход с пляжа" not in self.finreport_dict_beach:
            self.finreport_dict_beach["Выход с пляжа"] = 0, 0

    def agent_report(self):
        """Форминует отчет платежного агента в установленном формате."""

        self.agentreport_dict = {}
        self.agentreport_dict["Организация"] = [self.org1[0], self.org1[1]]
        for org, services in self.agent_dict.items():
            if org == "Не учитывать":
                continue

            self.agentreport_dict[org] = [0, 0]
            for serv in services:
                try:
                    if org == "Дата":
                        self.agentreport_dict[org][0] = self.itog_report_org1[serv][0]
                        self.agentreport_dict[org][1] = self.itog_report_org1[serv][1]
                    elif serv == "Депозит":
                        self.agentreport_dict[org][1] += self.itog_report_org1[serv][1]
                    elif serv == "Аквазона":
                        self.agentreport_dict[org][1] += self.itog_report_org1[serv][1]
                    elif serv == "Организация":
                        pass
                    else:
                        self.agentreport_dict[org][0] += self.itog_report_org1[serv][0]
                        self.agentreport_dict[org][1] += self.itog_report_org1[serv][1]
                except KeyError:
                    pass
                except TypeError:
                    pass

    def export_agent_report(self, agentreport_dict, date_from):
        """
        Сохраняет отчет платежного агента в виде Excel-файла в локальную директорию
        """
        # определяем стили
        h1 = Font(
            name="Times New Roman",
            size=18,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font = Font(
            name="Times New Roman",
            size=9,
            bold=False,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font_bold = Font(
            name="Times New Roman",
            size=9,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        fill = PatternFill(fill_type="solid", start_color="c1c1c1", end_color="c2c2c2")
        table_color = PatternFill(
            fill_type="solid", start_color="e2e2e2", end_color="e9e9e9"
        )
        align_top = Alignment(
            horizontal="general",
            vertical="top",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )
        border = Border(
            left=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
            top=Side(border_style="thin", color="FF000000"),
            bottom=Side(border_style="thin", color="FF000000"),
            diagonal=Side(border_style="thin", color="FF000000"),
            diagonal_direction=0,
            outline=Side(border_style="thin", color="FF000000"),
            vertical=Side(border_style="thin", color="FF000000"),
            horizontal=Side(border_style="thin", color="FF000000"),
        )
        align_left = Alignment(
            horizontal="left",
            vertical="bottom",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )

        column = ["", "A", "B", "C", "D", "E"]

        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Отчет платежного агента"
        # шрифты
        ws["A1"].font = h1
        # выравнивание
        ws["A1"].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 124
        ws.column_dimensions["B"].width = 1 / 7 * 80
        ws.column_dimensions["C"].width = 1 / 7 * 24
        ws.column_dimensions["D"].width = 1 / 7 * 175
        ws.column_dimensions["E"].width = 1 / 7 * 200

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = (
            "Отчет платежного агента по приему денежных средств"
        )
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        # шрифты
        ws[column[1] + self.row].font = h1
        # выравнивание
        ws[column[1] + self.row].alignment = align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f'{agentreport_dict["Организация"][1]}'
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top

        ws[column[1] + next_row()] = "За период с:"
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top
        ws[column[2] + self.row] = (agentreport_dict["Дата"][0]).strftime("%d.%m.%Y")
        ws[column[2] + self.row].font = font_bold
        ws[column[2] + self.row].alignment = align_top
        ws[column[3] + self.row] = "по"
        ws[column[3] + self.row].font = font
        ws[column[3] + self.row].alignment = align_top
        ws[column[4] + self.row] = (
            agentreport_dict["Дата"][1] - timedelta(1)
        ).strftime("%d.%m.%Y")
        ws[column[4] + self.row].font = font_bold
        ws[column[4] + self.row].alignment = align_top

        # ТАБЛИЦА
        self.color = False

        def merge_table():
            ws.merge_cells(
                start_row=self.row, start_column=1, end_row=self.row, end_column=4
            )
            ws[column[1] + self.row].font = font
            ws[column[5] + self.row].font = font
            ws[column[1] + self.row].alignment = align_top
            ws[column[5] + self.row].alignment = align_top
            ws[column[5] + self.row].number_format = "#,##0.00 ₽"
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1
            if self.color:
                b = 1
                while b < len(column):
                    ws[column[b] + self.row].fill = table_color
                    b += 1
                self.color = False
            else:
                self.color = True

        def merge_table_bold():
            ws.merge_cells(
                start_row=self.row, start_column=1, end_row=self.row, end_column=4
            )
            ws[column[1] + self.row].font = font_bold
            ws[column[5] + self.row].font = font_bold
            ws[column[1] + self.row].alignment = align_top
            ws[column[5] + self.row].alignment = align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        ws[column[1] + next_row()] = "Наименование поставщика услуг"
        ws[column[5] + self.row] = "Сумма"
        merge_table_bold()
        # раскрашивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = fill
            b += 1

        itog_sum = 0
        for line in agentreport_dict:
            if line != "Организация" and line != "Дата" and line != "ИТОГО":
                try:
                    itog_sum += agentreport_dict[line][1]
                    ws[column[1] + next_row()] = line
                    ws[column[5] + self.row] = agentreport_dict[line][1]
                    merge_table()
                except AttributeError:
                    pass

        ws[column[1] + next_row()] = "Итого"
        if itog_sum != agentreport_dict["ИТОГО"][1]:
            logger.error(
                f"Ошибка. Отчет платежного агента: сумма строк "
                f"({itog_sum}) не равна строке ИТОГО "
                f'({agentreport_dict["ИТОГО"][1]})'
            )
            logger.info(
                "Ошибка. Отчет платежного агента",
                "Ошибка. Отчет платежного агента: сумма строк "
                f"({itog_sum}) не равна строке ИТОГО "
                f'({agentreport_dict["ИТОГО"][1]})',
            )
        ws[column[5] + self.row] = itog_sum
        ws[column[5] + self.row].number_format = "#,##0.00 ₽"
        merge_table_bold()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if agentreport_dict["Дата"][0] == agentreport_dict["Дата"][1] - timedelta(1):
            date_ = datetime.strftime(agentreport_dict["Дата"][0], "%Y-%m-%d")
        else:
            date_ = (
                f'{datetime.strftime(agentreport_dict["Дата"][0], "%Y-%m-%d")} - '
                f'{datetime.strftime(agentreport_dict["Дата"][1], "%Y-%m-%d")}'
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f' Отчет платежного агента {agentreport_dict["Организация"][1]}'
            + ".xlsx"
        )
        logger.info(
            f"Сохранение отчета платежного агента "
            f'{agentreport_dict["Организация"][1]} в {path}'
        )
        path = self._yandex_repo.create_path(path, date_from)
        self._yandex_repo.save_file(path, wb)
        return path

    async def export_to_google_sheet(self, date_from, http_auth, googleservice):
        """
        Формирование и заполнение google-таблицы
        """
        logger.info("Сохранение Финансового отчета в Google-таблицах...")
        self.sheet_width = 35
        self.sheet2_width = 3
        self.sheet3_width = 14
        self.sheet4_width = 3
        self.sheet5_width = 3
        self.sheet6_width = 16
        self.sheet_height = 40
        self.sheet2_height = 40
        self.sheet4_height = 300
        self.sheet5_height = 300
        self.sheet6_height = 40

        self.data_report = datetime.strftime(self.finreport_dict["Дата"][0], "%m")
        month = [
            "",
            "Январь",
            "Февраль",
            "Март",
            "Апрель",
            "Май",
            "Июнь",
            "Июль",
            "Август",
            "Сентябрь",
            "Октябрь",
            "Ноябрь",
            "Декабрь",
        ]
        self.data_report = month[int(self.data_report)]

        doc_name = (
            f"{datetime.strftime(self.finreport_dict['Дата'][0], '%Y-%m')} "
            f"({self.data_report}) - Финансовый отчет по Аквапарку"
        )

        if (
            self.finreport_dict["Дата"][0] + timedelta(1)
            != self.finreport_dict["Дата"][1]
        ):
            logger.info("Экспорт отчета в Google Sheet за несколько дней невозможен!")
        else:
            google_report_id = (
                await self._report_config_service.get_financial_doc_id_by_date(
                    date_from
                )
            )
            if google_report_id is None:
                google_doc = create_new_google_doc(
                    googleservice=googleservice,
                    doc_name=doc_name,
                    data_report=self.data_report,
                    finreport_dict=self.finreport_dict,
                    http_auth=http_auth,
                    date_from=date_from,
                    sheet_width=self.sheet_width,
                    sheet2_width=self.sheet2_width,
                    sheet3_width=self.sheet3_width,
                    sheet4_width=self.sheet4_width,
                    sheet5_width=self.sheet5_width,
                    sheet6_width=self.sheet6_width,
                    sheet_height=self.sheet_height,
                    sheet2_height=self.sheet2_height,
                    sheet4_height=self.sheet4_height,
                    sheet5_height=self.sheet5_height,
                    sheet6_height=self.sheet6_height,
                )
                google_report_id = GoogleReportIdCreate(
                    month=google_doc[0],
                    doc_id=google_doc[1],
                    report_type="financial",
                    version=settings.google_api_settings.google_doc_version,
                )
                await self._report_config_service.add_google_report_id(google_report_id)
                logger.info(f"Создана новая таблица с Id: {google_report_id.doc_id}")

            if (
                google_report_id.version
                != settings.google_api_settings.google_doc_version
            ):
                error_message = (
                    f"Версия Финансового отчета ({google_report_id.version}) не соответствует текущей "
                    f"({settings.google_api_settings.google_doc_version}).\n"
                    f"Необходимо сначала удалить ссылку на старую версию, "
                    f"затем заново сформировать отчет с начала месяца."
                )
                logger.error(error_message)
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=error_message,
                )

            google_doc = (google_report_id.month, google_report_id.doc_id)
            self.spreadsheet = (
                googleservice.spreadsheets()
                .get(spreadsheetId=google_doc[1], ranges=[], includeGridData=True)
                .execute()
            )

            # -------------------------------- ЗАПОЛНЕНИЕ ДАННЫМИ ------------------------------------------------

            # Проверка нет ли текущей даты в таблице
            logger.info("Проверка нет ли текущей даты в таблице...")
            self.start_line = 1
            self.reprint = 2

            for line_table in self.spreadsheet["sheets"][0]["data"][0]["rowData"]:
                try:
                    if line_table["values"][0]["formattedValue"] == datetime.strftime(
                        self.finreport_dict["Дата"][0], "%d.%m.%Y"
                    ):
                        self.rewrite_google_sheet(googleservice)
                        self.reprint = 0
                        break
                    elif line_table["values"][0]["formattedValue"] == "ИТОГО":
                        break
                    else:
                        self.start_line += 1
                except KeyError:
                    self.start_line += 1
            if self.reprint:
                self.write_google_sheet(googleservice)
            # width_table = len(self.spreadsheet['sheets'][0]['data'][0]['rowData'][0]['values'])
        return True

    def rewrite_google_sheet(self, googleservice):
        """
        Заполнение google-таблицы в случае, если данные уже существуют
        """
        logger.warning("Перезапись уже существующей строки...")
        self.reprint = 1
        self.write_google_sheet(googleservice)

    def write_google_sheet(self, googleservice):
        """
        Заполнение google-таблицы
        """
        # SHEET 1
        logger.info("Заполнение листа 1...")
        sheetId = 0
        ss = Spreadsheet(
            self.spreadsheet["spreadsheetId"],
            sheetId,
            googleservice,
            self.spreadsheet["sheets"][sheetId]["properties"]["title"],
        )

        # Заполнение строки с данными
        weekday_rus = [
            "Понедельник",
            "Вторник",
            "Среда",
            "Четверг",
            "Пятница",
            "Суббота",
            "Воскресенье",
        ]
        self.nex_line = self.start_line

        control_total_sum = sum(
            [
                self.finreport_dict["Билеты аквапарка"][1],
                self.finreport_dict["Общепит"][1],
                self.finreport_dict["Билеты аквапарка КОРП"][1],
                self.finreport_dict["Прочее"][1],
                self.finreport_dict["Сопутствующие товары"][1],
                self.finreport_dict["Депозит"][1],
                self.finreport_dict["Штраф"][1],
                self.finreport_dict["Online Продажи"][1],
                self.finreport_dict["Фотоуслуги"][1],
            ]
        )

        if self.finreport_dict["ИТОГО"][1] != control_total_sum:
            logger.error("Несоответствие данных: Сумма услуг не равна итоговой сумме")
            logger.info(
                f"Несоответствие данных: Сумма услуг по группам + депозит ({control_total_sum}) "
                f"не равна итоговой сумме ({self.finreport_dict['ИТОГО'][1]}). \n"
                f"Рекомендуется проверить правильно ли разделены услуги по группам.",
            )

        ss.prepare_setValues(
            f"A{self.nex_line}:AI{self.nex_line}",
            [
                [
                    datetime.strftime(self.finreport_dict["Дата"][0], "%d.%m.%Y"),
                    weekday_rus[self.finreport_dict["Дата"][0].weekday()],
                    f"='План'!C{self.nex_line}",
                    f"{self.finreport_dict['Кол-во проходов'][0]}",
                    f"{self.finreport_dict_lastyear['Кол-во проходов'][0]}",
                    f"='План'!E{self.nex_line}",
                    f"={str(self.finreport_dict['ИТОГО'][1]).replace('.', ',')}"
                    f"-I{self.nex_line}+AG{self.nex_line}+AI{self.nex_line}+'Смайл'!C{self.nex_line}",
                    f"=IFERROR(G{self.nex_line}/D{self.nex_line};0)",
                    f"={str(self.finreport_dict['MaxBonus'][1]).replace('.', ',')}",
                    f"={str(self.finreport_dict_lastyear['ИТОГО'][1]).replace('.', ',')}"
                    f"-{str(self.finreport_dict_lastyear['MaxBonus'][1]).replace('.', ',')}"
                    f"+{str(self.finreport_dict_lastyear['Online Продажи'][1]).replace('.', ',')}",
                    self.finreport_dict["Билеты аквапарка"][0],
                    self.finreport_dict["Билеты аквапарка"][1],
                    f"=IFERROR(L{self.nex_line}/K{self.nex_line};0)",
                    self.finreport_dict["Депозит"][1],
                    self.finreport_dict["Штраф"][1],
                    f"='План'!I{self.nex_line}",
                    f"='План'!J{self.nex_line}",
                    f"=IFERROR(Q{self.nex_line}/P{self.nex_line};0)",
                    self.finreport_dict["Общепит"][0] + self.finreport_dict["Смайл"][0],
                    self.finreport_dict["Общепит"][1] + self.finreport_dict["Смайл"][1],
                    f"=IFERROR(T{self.nex_line}/S{self.nex_line};0)",
                    self.finreport_dict_lastyear["Общепит"][0]
                    + self.finreport_dict_lastyear["Смайл"][0],
                    self.finreport_dict_lastyear["Общепит"][1]
                    + self.finreport_dict_lastyear["Смайл"][1],
                    f"=IFERROR(W{self.nex_line}/V{self.nex_line};0)",
                    self.finreport_dict["Билеты аквапарка КОРП"][0],
                    self.finreport_dict["Билеты аквапарка КОРП"][1],
                    f"=IFERROR(Z{self.nex_line}/Y{self.nex_line};0)",
                    self.finreport_dict["Прочее"][0]
                    + self.finreport_dict["Сопутствующие товары"][0],
                    self.finreport_dict["Прочее"][1]
                    + self.finreport_dict["Сопутствующие товары"][1],
                    self.finreport_dict["Online Продажи"][0],
                    self.finreport_dict["Online Продажи"][1],
                    f"=IFERROR(AE{self.nex_line}/AD{self.nex_line};0)",
                    0,
                    self.finreport_dict["Фотоуслуги"][1],
                    0,
                ]
            ],
            "ROWS",
        )

        # Задание форматы вывода строки
        ss.prepare_setCellsFormats(
            f"A{self.nex_line}:AI{self.nex_line}",
            [
                [
                    {
                        "numberFormat": {"type": "DATE", "pattern": "dd.mm.yyyy"},
                        "horizontalAlignment": "LEFT",
                    },
                    {"numberFormat": {}},
                    {"numberFormat": {}},
                    {"numberFormat": {}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                ]
            ],
        )
        # Цвет фона ячеек
        if self.nex_line % 2 != 0:
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:AI{self.nex_line}",
                {"backgroundColor": functions.htmlColorToJSON("#fef8e3")},
                fields="userEnteredFormat.backgroundColor",
            )

        # Бордер
        for j in range(self.sheet_width):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
        ss.runPrepared()

        # ------------------------------------------- Заполнение ИТОГО --------------------------------------
        # Вычисление последней строки в таблице
        logger.info("Заполнение строки ИТОГО на листе 1...")

        self.sheet2_line = 1
        for line_table in self.spreadsheet["sheets"][2]["data"][0]["rowData"]:
            try:
                if line_table["values"][0]["formattedValue"] == "ИТОГО":
                    break
                else:
                    self.sheet2_line += 1
            except KeyError:
                self.sheet2_line += 1

        for i, line_table in enumerate(
            self.spreadsheet["sheets"][0]["data"][0]["rowData"]
        ):
            try:
                if line_table["values"][0]["formattedValue"] == "ИТОГО":
                    # Если строка переписывается - итого на 1 поз вниз, если новая - на 2 поз
                    height_table = i + self.reprint
                    break
                else:
                    height_table = 4
            except KeyError:
                pass

        ss.prepare_setValues(
            f"A{height_table}:AI{height_table}",
            [
                [
                    "ИТОГО",
                    "",
                    f"=SUM(C3:C{height_table - 1})",
                    f"=SUM(D3:D{height_table - 1})",
                    f"=SUM(E3:E{height_table - 1})",
                    f"=SUM(F3:F{height_table - 1})",
                    f"=SUM(G3:G{height_table - 1})",
                    f"=IFERROR(ROUND(G{height_table}/D{height_table};2);0)",
                    f"=SUM(I3:I{height_table - 1})",
                    f"=SUM(J3:J{height_table - 1})",
                    f"=SUM(K3:K{height_table - 1})",
                    f"=SUM(L3:L{height_table - 1})",
                    f"=IFERROR(ROUND(L{height_table}/K{height_table};2);0)",
                    f"=SUM(N3:N{height_table - 1})",
                    f"=SUM(O3:O{height_table - 1})",
                    f"=SUM(P3:P{height_table - 1})",
                    f"=SUM(Q3:Q{height_table - 1})",
                    f"=IFERROR(ROUND(Q{height_table}/P{height_table};2);0)",
                    f"=SUM(S3:S{height_table - 1})",
                    f"=SUM(T3:T{height_table - 1})",
                    f"=IFERROR(ROUND(T{height_table}/S{height_table};2);0)",
                    f"=SUM(V3:V{height_table - 1})",
                    f"=SUM(W3:W{height_table - 1})",
                    f"=IFERROR(ROUND(W{height_table}/V{height_table};2);0)",
                    f"=SUM(Y3:Y{height_table - 1})",
                    f"=SUM(Z3:Z{height_table - 1})",
                    f"=IFERROR(ROUND(Z{height_table}/Y{height_table};2);0)",
                    f"=SUM(AB3:AB{height_table - 1})",
                    f"=SUM(AC3:AC{height_table - 1})",
                    f"=SUM(AD3:AD{height_table - 1})",
                    f"=SUM(AE3:AE{height_table - 1})",
                    f"=IFERROR(ROUND(AE{height_table}/AD{height_table};2);0)",
                    f"=SUM(AG3:AG{height_table - 1})",
                    f"=SUM(AH3:AH{height_table - 1})",
                    f"=SUM(AI3:AI{height_table - 1})",
                ]
            ],
            "ROWS",
        )
        ss.prepare_setValues(
            f"A{height_table + 1}:D{height_table + 1}",
            [
                [
                    "Выполнение плана (трафик)",
                    "",
                    f"=IFERROR('План'!C{self.sheet2_line};0)",
                    f"=IFERROR(ROUND(D{height_table}/C{height_table + 1};2);0)",
                ]
            ],
            "ROWS",
        )
        ss.prepare_setValues(
            f"A{height_table + 2}:D{height_table + 2}",
            [
                [
                    "Выполнение плана (доход)",
                    "",
                    f"=IFERROR('План'!E{self.sheet2_line};0)",
                    f"=IFERROR(ROUND(G{height_table}/C{height_table + 2};2);0)",
                ]
            ],
            "ROWS",
        )

        # Задание формата вывода строки
        ss.prepare_setCellsFormats(
            f"A{height_table}:AI{height_table}",
            [
                [
                    {"textFormat": {"bold": True}},
                    {"textFormat": {"bold": True}},
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                ]
            ],
        )
        ss.prepare_setCellsFormats(
            f"A{height_table + 1}:D{height_table + 1}",
            [
                [
                    {"textFormat": {"bold": True}},
                    {"textFormat": {"bold": True}},
                    {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "RIGHT",
                        "numberFormat": {},
                    },
                    {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "RIGHT",
                        "numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00%"},
                    },
                ]
            ],
        )
        ss.prepare_setCellsFormats(
            f"A{height_table + 2}:D{height_table + 2}",
            [
                [
                    {"textFormat": {"bold": True}},
                    {"textFormat": {"bold": True}},
                    {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "RIGHT",
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                    },
                    {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "RIGHT",
                        "numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00%"},
                    },
                ]
            ],
        )

        # Цвет фона ячеек
        ss.prepare_setCellsFormat(
            f"A{height_table}:AI{height_table}",
            {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
            fields="userEnteredFormat.backgroundColor",
        )
        ss.prepare_setCellsFormat(
            f"A{height_table + 1}:D{height_table + 1}",
            {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
            fields="userEnteredFormat.backgroundColor",
        )
        ss.prepare_setCellsFormat(
            f"A{height_table + 2}:D{height_table + 2}",
            {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
            fields="userEnteredFormat.backgroundColor",
        )

        # Бордер
        for j in range(self.sheet_width):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
        for j in range(4):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table,
                            "endRowIndex": height_table + 1,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table,
                            "endRowIndex": height_table + 1,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table,
                            "endRowIndex": height_table + 1,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table,
                            "endRowIndex": height_table + 1,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
        for j in range(4):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table + 1,
                            "endRowIndex": height_table + 2,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table + 1,
                            "endRowIndex": height_table + 2,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table + 1,
                            "endRowIndex": height_table + 2,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table + 1,
                            "endRowIndex": height_table + 2,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
        ss.runPrepared()

        logger.info("Заполнение листа 2...")
        sheetId = 1
        ss = Spreadsheet(
            self.spreadsheet["spreadsheetId"],
            sheetId,
            googleservice,
            self.spreadsheet["sheets"][sheetId]["properties"]["title"],
        )

        self.nex_line = self.start_line

        ss.prepare_setValues(
            f"A{self.nex_line}:C{self.nex_line}",
            [
                [
                    datetime.strftime(self.finreport_dict["Дата"][0], "%d.%m.%Y"),
                    self.smile_report.total_count,
                    self.smile_report.total_sum,
                ]
            ],
            "ROWS",
        )

        # Задание форматы вывода строки
        ss.prepare_setCellsFormats(
            f"A{self.nex_line}:C{self.nex_line}",
            [
                [
                    {
                        "numberFormat": {"type": "DATE", "pattern": "dd.mm.yyyy"},
                        "horizontalAlignment": "LEFT",
                    },
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"}},
                ]
            ],
        )
        # Цвет фона ячеек
        if self.nex_line % 2 != 0:
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {"backgroundColor": functions.htmlColorToJSON("#fef8e3")},
                fields="userEnteredFormat.backgroundColor",
            )

        # Бордер
        for j in range(self.sheet2_width):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )

        # ------------------------------------------- Заполнение ИТОГО --------------------------------------
        # Вычисление последней строки в таблице
        logger.info("Заполнение строки ИТОГО на листе 2...")

        for i, line_table in enumerate(
            self.spreadsheet["sheets"][1]["data"][0]["rowData"]
        ):
            try:
                if line_table["values"][0]["formattedValue"] == "ИТОГО":
                    # Если строка переписывается - итого на 1 поз вниз, если новая - на 2 поз
                    height_table = i + self.reprint
                    break
                else:
                    height_table = 4
            except KeyError:
                pass

        ss.prepare_setValues(
            f"A{height_table}:C{height_table}",
            [
                [
                    "ИТОГО",
                    f"=SUM(B3:B{height_table - 1})",
                    f"=SUM(C3:C{height_table - 1})",
                ]
            ],
            "ROWS",
        )

        # Задание формата вывода строки
        ss.prepare_setCellsFormats(
            f"A{height_table}:C{height_table}",
            [
                [
                    {"textFormat": {"bold": True}},
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {"type": "CURRENCY", "pattern": "#0[$ ₽]"},
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                ]
            ],
        )

        # Цвет фона ячеек
        ss.prepare_setCellsFormat(
            f"A{height_table}:C{height_table}",
            {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
            fields="userEnteredFormat.backgroundColor",
        )

        # Бордер
        for j in range(self.sheet2_width):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
        ss.runPrepared()

        if self.itog_report_month:
            # SHEET 4
            logger.info("Заполнение  листа 4...")
            sheetId = 3
            ss = Spreadsheet(
                self.spreadsheet["spreadsheetId"],
                sheetId,
                googleservice,
                self.spreadsheet["sheets"][sheetId]["properties"]["title"],
            )

            self.nex_line = 1
            ss.prepare_setValues(
                f"A{self.nex_line}:C{self.nex_line}",
                [["Итоговый отчет", "", ""]],
                "ROWS",
            )
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {
                    "horizontalAlignment": "LEFT",
                    "textFormat": {"bold": True, "fontSize": 18},
                },
            )

            self.nex_line += 1
            ss.prepare_setValues(
                f"A{self.nex_line}:C{self.nex_line}",
                [
                    [
                        f"За {self.data_report} {datetime.strftime(self.finreport_dict['Дата'][0], '%Y')}",
                        "",
                        "",
                    ]
                ],
                "ROWS",
            )
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {"horizontalAlignment": "LEFT", "textFormat": {"bold": False}},
            )

            self.nex_line += 2
            ss.prepare_setValues(
                f"A{self.nex_line}:C{self.nex_line}",
                [["Название", "Количество", "Сумма"]],
                "ROWS",
            )
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {
                    "horizontalAlignment": "LEFT",
                    "textFormat": {"bold": True, "fontSize": 14},
                },
            )
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {"backgroundColor": functions.htmlColorToJSON("#f7cb4d")},
                fields="userEnteredFormat.backgroundColor",
            )
            for j in range(self.sheet4_width):
                ss.requests.append(
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": ss.sheetId,
                                "startRowIndex": self.nex_line - 1,
                                "endRowIndex": self.nex_line,
                                "startColumnIndex": j,
                                "endColumnIndex": j + 1,
                            },
                            "top": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {"red": 0, "green": 0, "blue": 0},
                            },
                        }
                    }
                )
                ss.requests.append(
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": ss.sheetId,
                                "startRowIndex": self.nex_line - 1,
                                "endRowIndex": self.nex_line,
                                "startColumnIndex": j,
                                "endColumnIndex": j + 1,
                            },
                            "right": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {
                                    "red": 0,
                                    "green": 0,
                                    "blue": 0,
                                    "alpha": 1.0,
                                },
                            },
                        }
                    }
                )
                ss.requests.append(
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": ss.sheetId,
                                "startRowIndex": self.nex_line - 1,
                                "endRowIndex": self.nex_line,
                                "startColumnIndex": j,
                                "endColumnIndex": j + 1,
                            },
                            "left": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {
                                    "red": 0,
                                    "green": 0,
                                    "blue": 0,
                                    "alpha": 1.0,
                                },
                            },
                        }
                    }
                )
                ss.requests.append(
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": ss.sheetId,
                                "startRowIndex": self.nex_line - 1,
                                "endRowIndex": self.nex_line,
                                "startColumnIndex": j,
                                "endColumnIndex": j + 1,
                            },
                            "bottom": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {
                                    "red": 0,
                                    "green": 0,
                                    "blue": 0,
                                    "alpha": 1.0,
                                },
                            },
                        }
                    }
                )
            ss.runPrepared()

            for group, group_values in self.finreport_dict_month.items():
                if group == "Контрольная сумма":
                    continue
                if group == "Дата":
                    continue
                self.nex_line += 1
                ss.prepare_setValues(
                    f"A{self.nex_line}:C{self.nex_line}",
                    [
                        [
                            group,
                            group_values["Итого по группе"][0][1],
                            group_values["Итого по группе"][0][2],
                        ]
                    ],
                    "ROWS",
                )
                ss.prepare_setCellsFormats(
                    f"A{self.nex_line}:C{self.nex_line}",
                    [
                        [
                            {"textFormat": {"bold": True, "fontSize": 12}},
                            {
                                "textFormat": {"bold": True, "fontSize": 12},
                                "horizontalAlignment": "RIGHT",
                                "numberFormat": {},
                            },
                            {
                                "textFormat": {"bold": True, "fontSize": 12},
                                "horizontalAlignment": "RIGHT",
                                "numberFormat": {
                                    "type": "CURRENCY",
                                    "pattern": "#,##0.00[$ ₽]",
                                },
                            },
                        ]
                    ],
                )
                ss.prepare_setCellsFormat(
                    f"A{self.nex_line}:C{self.nex_line}",
                    {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
                    fields="userEnteredFormat.backgroundColor",
                )
                for j in range(self.sheet4_width):
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "top": {
                                    "style": "SOLID",
                                    "width": 1,
                                    "color": {"red": 0, "green": 0, "blue": 0},
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "right": {
                                    "style": "SOLID",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "left": {
                                    "style": "SOLID",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "bottom": {
                                    "style": "SOLID",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                for folder, folder_values in group_values.items():
                    if folder == "Итого по группе":
                        continue
                    if folder == "":
                        continue
                    self.nex_line += 1
                    if folder is None:
                        folder_name = "Без группировки"
                    else:
                        folder_name = folder
                    ss.prepare_setValues(
                        f"A{self.nex_line}:C{self.nex_line}",
                        [
                            [
                                folder_name,
                                folder_values[0][1],
                                folder_values[0][2],
                            ]
                        ],
                        "ROWS",
                    )
                    ss.prepare_setCellsFormats(
                        f"A{self.nex_line}:C{self.nex_line}",
                        [
                            [
                                {"textFormat": {"bold": True}},
                                {
                                    "textFormat": {"bold": True},
                                    "horizontalAlignment": "RIGHT",
                                    "numberFormat": {},
                                },
                                {
                                    "textFormat": {"bold": True},
                                    "horizontalAlignment": "RIGHT",
                                    "numberFormat": {
                                        "type": "CURRENCY",
                                        "pattern": "#,##0.00[$ ₽]",
                                    },
                                },
                            ]
                        ],
                    )
                    ss.prepare_setCellsFormat(
                        f"A{self.nex_line}:C{self.nex_line}",
                        {"backgroundColor": functions.htmlColorToJSON("#fef8e3")},
                        fields="userEnteredFormat.backgroundColor",
                    )
                    for j in range(self.sheet4_width):
                        ss.requests.append(
                            {
                                "updateBorders": {
                                    "range": {
                                        "sheetId": ss.sheetId,
                                        "startRowIndex": self.nex_line - 1,
                                        "endRowIndex": self.nex_line,
                                        "startColumnIndex": j,
                                        "endColumnIndex": j + 1,
                                    },
                                    "top": {
                                        "style": "SOLID",
                                        "width": 1,
                                        "color": {"red": 0, "green": 0, "blue": 0},
                                    },
                                }
                            }
                        )
                        ss.requests.append(
                            {
                                "updateBorders": {
                                    "range": {
                                        "sheetId": ss.sheetId,
                                        "startRowIndex": self.nex_line - 1,
                                        "endRowIndex": self.nex_line,
                                        "startColumnIndex": j,
                                        "endColumnIndex": j + 1,
                                    },
                                    "right": {
                                        "style": "SOLID",
                                        "width": 1,
                                        "color": {
                                            "red": 0,
                                            "green": 0,
                                            "blue": 0,
                                            "alpha": 1.0,
                                        },
                                    },
                                }
                            }
                        )
                        ss.requests.append(
                            {
                                "updateBorders": {
                                    "range": {
                                        "sheetId": ss.sheetId,
                                        "startRowIndex": self.nex_line - 1,
                                        "endRowIndex": self.nex_line,
                                        "startColumnIndex": j,
                                        "endColumnIndex": j + 1,
                                    },
                                    "left": {
                                        "style": "SOLID",
                                        "width": 1,
                                        "color": {
                                            "red": 0,
                                            "green": 0,
                                            "blue": 0,
                                            "alpha": 1.0,
                                        },
                                    },
                                }
                            }
                        )
                        ss.requests.append(
                            {
                                "updateBorders": {
                                    "range": {
                                        "sheetId": ss.sheetId,
                                        "startRowIndex": self.nex_line - 1,
                                        "endRowIndex": self.nex_line,
                                        "startColumnIndex": j,
                                        "endColumnIndex": j + 1,
                                    },
                                    "bottom": {
                                        "style": "SOLID",
                                        "width": 1,
                                        "color": {
                                            "red": 0,
                                            "green": 0,
                                            "blue": 0,
                                            "alpha": 1.0,
                                        },
                                    },
                                }
                            }
                        )
                    for service_name, service_count, service_sum in folder_values:
                        if service_name == "Итого по папке":
                            continue
                        self.nex_line += 1
                        ss.prepare_setValues(
                            f"A{self.nex_line}:C{self.nex_line}",
                            [[service_name, service_count, service_sum]],
                            "ROWS",
                        )
                        ss.prepare_setCellsFormats(
                            f"A{self.nex_line}:C{self.nex_line}",
                            [
                                [
                                    {"textFormat": {"bold": False}},
                                    {
                                        "textFormat": {"bold": False},
                                        "horizontalAlignment": "RIGHT",
                                        "numberFormat": {},
                                    },
                                    {
                                        "textFormat": {"bold": False},
                                        "horizontalAlignment": "RIGHT",
                                        "numberFormat": {
                                            "type": "CURRENCY",
                                            "pattern": "#,##0.00[$ ₽]",
                                        },
                                    },
                                ]
                            ],
                        )
                        for j in range(self.sheet4_width):
                            ss.requests.append(
                                {
                                    "updateBorders": {
                                        "range": {
                                            "sheetId": ss.sheetId,
                                            "startRowIndex": self.nex_line - 1,
                                            "endRowIndex": self.nex_line,
                                            "startColumnIndex": j,
                                            "endColumnIndex": j + 1,
                                        },
                                        "top": {
                                            "style": "SOLID",
                                            "width": 1,
                                            "color": {"red": 0, "green": 0, "blue": 0},
                                        },
                                    }
                                }
                            )
                            ss.requests.append(
                                {
                                    "updateBorders": {
                                        "range": {
                                            "sheetId": ss.sheetId,
                                            "startRowIndex": self.nex_line - 1,
                                            "endRowIndex": self.nex_line,
                                            "startColumnIndex": j,
                                            "endColumnIndex": j + 1,
                                        },
                                        "right": {
                                            "style": "SOLID",
                                            "width": 1,
                                            "color": {
                                                "red": 0,
                                                "green": 0,
                                                "blue": 0,
                                                "alpha": 1.0,
                                            },
                                        },
                                    }
                                }
                            )
                            ss.requests.append(
                                {
                                    "updateBorders": {
                                        "range": {
                                            "sheetId": ss.sheetId,
                                            "startRowIndex": self.nex_line - 1,
                                            "endRowIndex": self.nex_line,
                                            "startColumnIndex": j,
                                            "endColumnIndex": j + 1,
                                        },
                                        "left": {
                                            "style": "SOLID",
                                            "width": 1,
                                            "color": {
                                                "red": 0,
                                                "green": 0,
                                                "blue": 0,
                                                "alpha": 1.0,
                                            },
                                        },
                                    }
                                }
                            )
                            ss.requests.append(
                                {
                                    "updateBorders": {
                                        "range": {
                                            "sheetId": ss.sheetId,
                                            "startRowIndex": self.nex_line - 1,
                                            "endRowIndex": self.nex_line,
                                            "startColumnIndex": j,
                                            "endColumnIndex": j + 1,
                                        },
                                        "bottom": {
                                            "style": "SOLID",
                                            "width": 1,
                                            "color": {
                                                "red": 0,
                                                "green": 0,
                                                "blue": 0,
                                                "alpha": 1.0,
                                            },
                                        },
                                    }
                                }
                            )

            while self.nex_line < self.sheet4_height:
                self.nex_line += 1
                ss.prepare_setValues(
                    f"A{self.nex_line}:C{self.nex_line}", [["", "", ""]], "ROWS"
                )
                ss.prepare_setCellsFormat(
                    f"A{self.nex_line}:C{self.nex_line}",
                    {
                        "horizontalAlignment": "LEFT",
                        "textFormat": {"bold": False, "fontSize": 10},
                    },
                )
                ss.prepare_setCellsFormat(
                    f"A{self.nex_line}:C{self.nex_line}",
                    {"backgroundColor": functions.htmlColorToJSON("#ffffff")},
                    fields="userEnteredFormat.backgroundColor",
                )
                for j in range(self.sheet4_width):
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "right": {
                                    "style": "NONE",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "left": {
                                    "style": "NONE",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "bottom": {
                                    "style": "NONE",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
            ss.runPrepared()

            # SHEET 4
            logger.info("Заполнение  листа 5...")
            sheetId = 4
            ss = Spreadsheet(
                self.spreadsheet["spreadsheetId"],
                sheetId,
                googleservice,
                self.spreadsheet["sheets"][sheetId]["properties"]["title"],
            )

            self.nex_line = 1
            ss.prepare_setValues(
                f"A{self.nex_line}:C{self.nex_line}",
                [["Итоговый отчет платежного агента", "", ""]],
                "ROWS",
            )
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {
                    "horizontalAlignment": "LEFT",
                    "textFormat": {"bold": True, "fontSize": 18},
                },
            )

            self.nex_line += 1
            ss.prepare_setValues(
                f"A{self.nex_line}:C{self.nex_line}",
                [
                    [
                        f"За {self.data_report} {datetime.strftime(self.finreport_dict['Дата'][0], '%Y')}",
                        "",
                        "",
                    ]
                ],
                "ROWS",
            )
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {"horizontalAlignment": "LEFT", "textFormat": {"bold": False}},
            )

            self.nex_line += 2
            ss.prepare_setValues(
                f"A{self.nex_line}:C{self.nex_line}",
                [["Название", "Количество", "Сумма"]],
                "ROWS",
            )
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {
                    "horizontalAlignment": "LEFT",
                    "textFormat": {"bold": True, "fontSize": 14},
                },
            )
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:C{self.nex_line}",
                {"backgroundColor": functions.htmlColorToJSON("#f7cb4d")},
                fields="userEnteredFormat.backgroundColor",
            )
            for j in range(self.sheet5_width):
                ss.requests.append(
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": ss.sheetId,
                                "startRowIndex": self.nex_line - 1,
                                "endRowIndex": self.nex_line,
                                "startColumnIndex": j,
                                "endColumnIndex": j + 1,
                            },
                            "top": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {"red": 0, "green": 0, "blue": 0},
                            },
                        }
                    }
                )
                ss.requests.append(
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": ss.sheetId,
                                "startRowIndex": self.nex_line - 1,
                                "endRowIndex": self.nex_line,
                                "startColumnIndex": j,
                                "endColumnIndex": j + 1,
                            },
                            "right": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {
                                    "red": 0,
                                    "green": 0,
                                    "blue": 0,
                                    "alpha": 1.0,
                                },
                            },
                        }
                    }
                )
                ss.requests.append(
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": ss.sheetId,
                                "startRowIndex": self.nex_line - 1,
                                "endRowIndex": self.nex_line,
                                "startColumnIndex": j,
                                "endColumnIndex": j + 1,
                            },
                            "left": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {
                                    "red": 0,
                                    "green": 0,
                                    "blue": 0,
                                    "alpha": 1.0,
                                },
                            },
                        }
                    }
                )
                ss.requests.append(
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": ss.sheetId,
                                "startRowIndex": self.nex_line - 1,
                                "endRowIndex": self.nex_line,
                                "startColumnIndex": j,
                                "endColumnIndex": j + 1,
                            },
                            "bottom": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {
                                    "red": 0,
                                    "green": 0,
                                    "blue": 0,
                                    "alpha": 1.0,
                                },
                            },
                        }
                    }
                )
            ss.runPrepared()

            for group in self.agentreport_dict_month:
                if group == "Контрольная сумма":
                    continue
                if group == "Дата":
                    continue
                if group == "Не учитывать":
                    continue
                self.nex_line += 1
                ss.prepare_setValues(
                    f"A{self.nex_line}:C{self.nex_line}",
                    [
                        [
                            group,
                            self.agentreport_dict_month[group]["Итого по группе"][0][1],
                            self.agentreport_dict_month[group]["Итого по группе"][0][2],
                        ]
                    ],
                    "ROWS",
                )
                ss.prepare_setCellsFormats(
                    f"A{self.nex_line}:C{self.nex_line}",
                    [
                        [
                            {"textFormat": {"bold": True, "fontSize": 12}},
                            {
                                "textFormat": {"bold": True, "fontSize": 12},
                                "horizontalAlignment": "RIGHT",
                                "numberFormat": {},
                            },
                            {
                                "textFormat": {"bold": True, "fontSize": 12},
                                "horizontalAlignment": "RIGHT",
                                "numberFormat": {
                                    "type": "CURRENCY",
                                    "pattern": "#,##0.00[$ ₽]",
                                },
                            },
                        ]
                    ],
                )
                ss.prepare_setCellsFormat(
                    f"A{self.nex_line}:C{self.nex_line}",
                    {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
                    fields="userEnteredFormat.backgroundColor",
                )
                for j in range(self.sheet4_width):
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "top": {
                                    "style": "SOLID",
                                    "width": 1,
                                    "color": {"red": 0, "green": 0, "blue": 0},
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "right": {
                                    "style": "SOLID",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "left": {
                                    "style": "SOLID",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "bottom": {
                                    "style": "SOLID",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                for folder in self.agentreport_dict_month[group]:
                    if folder == "Итого по группе":
                        continue
                    if folder == "":
                        continue
                    self.nex_line += 1
                    if folder is None:
                        folder_name = "Без группировки"
                    else:
                        folder_name = folder
                    ss.prepare_setValues(
                        f"A{self.nex_line}:C{self.nex_line}",
                        [
                            [
                                folder_name,
                                self.agentreport_dict_month[group][folder][0][1],
                                self.agentreport_dict_month[group][folder][0][2],
                            ]
                        ],
                        "ROWS",
                    )
                    ss.prepare_setCellsFormats(
                        f"A{self.nex_line}:C{self.nex_line}",
                        [
                            [
                                {"textFormat": {"bold": True}},
                                {
                                    "textFormat": {"bold": True},
                                    "horizontalAlignment": "RIGHT",
                                    "numberFormat": {},
                                },
                                {
                                    "textFormat": {"bold": True},
                                    "horizontalAlignment": "RIGHT",
                                    "numberFormat": {
                                        "type": "CURRENCY",
                                        "pattern": "#,##0.00[$ ₽]",
                                    },
                                },
                            ]
                        ],
                    )
                    ss.prepare_setCellsFormat(
                        f"A{self.nex_line}:C{self.nex_line}",
                        {"backgroundColor": functions.htmlColorToJSON("#fef8e3")},
                        fields="userEnteredFormat.backgroundColor",
                    )
                    for j in range(self.sheet4_width):
                        ss.requests.append(
                            {
                                "updateBorders": {
                                    "range": {
                                        "sheetId": ss.sheetId,
                                        "startRowIndex": self.nex_line - 1,
                                        "endRowIndex": self.nex_line,
                                        "startColumnIndex": j,
                                        "endColumnIndex": j + 1,
                                    },
                                    "top": {
                                        "style": "SOLID",
                                        "width": 1,
                                        "color": {"red": 0, "green": 0, "blue": 0},
                                    },
                                }
                            }
                        )
                        ss.requests.append(
                            {
                                "updateBorders": {
                                    "range": {
                                        "sheetId": ss.sheetId,
                                        "startRowIndex": self.nex_line - 1,
                                        "endRowIndex": self.nex_line,
                                        "startColumnIndex": j,
                                        "endColumnIndex": j + 1,
                                    },
                                    "right": {
                                        "style": "SOLID",
                                        "width": 1,
                                        "color": {
                                            "red": 0,
                                            "green": 0,
                                            "blue": 0,
                                            "alpha": 1.0,
                                        },
                                    },
                                }
                            }
                        )
                        ss.requests.append(
                            {
                                "updateBorders": {
                                    "range": {
                                        "sheetId": ss.sheetId,
                                        "startRowIndex": self.nex_line - 1,
                                        "endRowIndex": self.nex_line,
                                        "startColumnIndex": j,
                                        "endColumnIndex": j + 1,
                                    },
                                    "left": {
                                        "style": "SOLID",
                                        "width": 1,
                                        "color": {
                                            "red": 0,
                                            "green": 0,
                                            "blue": 0,
                                            "alpha": 1.0,
                                        },
                                    },
                                }
                            }
                        )
                        ss.requests.append(
                            {
                                "updateBorders": {
                                    "range": {
                                        "sheetId": ss.sheetId,
                                        "startRowIndex": self.nex_line - 1,
                                        "endRowIndex": self.nex_line,
                                        "startColumnIndex": j,
                                        "endColumnIndex": j + 1,
                                    },
                                    "bottom": {
                                        "style": "SOLID",
                                        "width": 1,
                                        "color": {
                                            "red": 0,
                                            "green": 0,
                                            "blue": 0,
                                            "alpha": 1.0,
                                        },
                                    },
                                }
                            }
                        )
                    for servise in self.agentreport_dict_month[group][folder]:
                        if servise[0] == "Итого по папке":
                            continue
                        self.nex_line += 1
                        ss.prepare_setValues(
                            f"A{self.nex_line}:C{self.nex_line}",
                            [[servise[0], servise[1], servise[2]]],
                            "ROWS",
                        )
                        ss.prepare_setCellsFormats(
                            f"A{self.nex_line}:C{self.nex_line}",
                            [
                                [
                                    {"textFormat": {"bold": False}},
                                    {
                                        "textFormat": {"bold": False},
                                        "horizontalAlignment": "RIGHT",
                                        "numberFormat": {},
                                    },
                                    {
                                        "textFormat": {"bold": False},
                                        "horizontalAlignment": "RIGHT",
                                        "numberFormat": {
                                            "type": "CURRENCY",
                                            "pattern": "#,##0.00[$ ₽]",
                                        },
                                    },
                                ]
                            ],
                        )
                        for j in range(self.sheet5_width):
                            ss.requests.append(
                                {
                                    "updateBorders": {
                                        "range": {
                                            "sheetId": ss.sheetId,
                                            "startRowIndex": self.nex_line - 1,
                                            "endRowIndex": self.nex_line,
                                            "startColumnIndex": j,
                                            "endColumnIndex": j + 1,
                                        },
                                        "top": {
                                            "style": "SOLID",
                                            "width": 1,
                                            "color": {"red": 0, "green": 0, "blue": 0},
                                        },
                                    }
                                }
                            )
                            ss.requests.append(
                                {
                                    "updateBorders": {
                                        "range": {
                                            "sheetId": ss.sheetId,
                                            "startRowIndex": self.nex_line - 1,
                                            "endRowIndex": self.nex_line,
                                            "startColumnIndex": j,
                                            "endColumnIndex": j + 1,
                                        },
                                        "right": {
                                            "style": "SOLID",
                                            "width": 1,
                                            "color": {
                                                "red": 0,
                                                "green": 0,
                                                "blue": 0,
                                                "alpha": 1.0,
                                            },
                                        },
                                    }
                                }
                            )
                            ss.requests.append(
                                {
                                    "updateBorders": {
                                        "range": {
                                            "sheetId": ss.sheetId,
                                            "startRowIndex": self.nex_line - 1,
                                            "endRowIndex": self.nex_line,
                                            "startColumnIndex": j,
                                            "endColumnIndex": j + 1,
                                        },
                                        "left": {
                                            "style": "SOLID",
                                            "width": 1,
                                            "color": {
                                                "red": 0,
                                                "green": 0,
                                                "blue": 0,
                                                "alpha": 1.0,
                                            },
                                        },
                                    }
                                }
                            )
                            ss.requests.append(
                                {
                                    "updateBorders": {
                                        "range": {
                                            "sheetId": ss.sheetId,
                                            "startRowIndex": self.nex_line - 1,
                                            "endRowIndex": self.nex_line,
                                            "startColumnIndex": j,
                                            "endColumnIndex": j + 1,
                                        },
                                        "bottom": {
                                            "style": "SOLID",
                                            "width": 1,
                                            "color": {
                                                "red": 0,
                                                "green": 0,
                                                "blue": 0,
                                                "alpha": 1.0,
                                            },
                                        },
                                    }
                                }
                            )

            while self.nex_line < self.sheet5_height:
                self.nex_line += 1
                ss.prepare_setValues(
                    f"A{self.nex_line}:C{self.nex_line}", [["", "", ""]], "ROWS"
                )
                ss.prepare_setCellsFormat(
                    f"A{self.nex_line}:C{self.nex_line}",
                    {
                        "horizontalAlignment": "LEFT",
                        "textFormat": {"bold": False, "fontSize": 10},
                    },
                )
                ss.prepare_setCellsFormat(
                    f"A{self.nex_line}:C{self.nex_line}",
                    {"backgroundColor": functions.htmlColorToJSON("#ffffff")},
                    fields="userEnteredFormat.backgroundColor",
                )
                for j in range(self.sheet5_width):
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "right": {
                                    "style": "NONE",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "left": {
                                    "style": "NONE",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
                    ss.requests.append(
                        {
                            "updateBorders": {
                                "range": {
                                    "sheetId": ss.sheetId,
                                    "startRowIndex": self.nex_line - 1,
                                    "endRowIndex": self.nex_line,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "bottom": {
                                    "style": "NONE",
                                    "width": 1,
                                    "color": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0,
                                        "alpha": 1.0,
                                    },
                                },
                            }
                        }
                    )
            ss.runPrepared()

        # Заполнение листа 6
        logger.info("Заполнение листа 6...")
        sheetId = 5
        ss = Spreadsheet(
            self.spreadsheet["spreadsheetId"],
            sheetId,
            googleservice,
            self.spreadsheet["sheets"][sheetId]["properties"]["title"],
        )

        # Заполнение строки с данными
        weekday_rus = [
            "Понедельник",
            "Вторник",
            "Среда",
            "Четверг",
            "Пятница",
            "Суббота",
            "Воскресенье",
        ]
        self.nex_line = self.start_line
        ss.prepare_setValues(
            f"A{self.nex_line}:P{self.nex_line}",
            [
                [
                    datetime.strftime(self.finreport_dict_beach["Дата"][0], "%d.%m.%Y"),
                    weekday_rus[self.finreport_dict_beach["Дата"][0].weekday()],
                    f"='План'!L{self.nex_line}",
                    self.finreport_dict_beach["Выход с пляжа"][0],
                    f"='План'!M{self.nex_line}",
                    str(self.finreport_dict_beach["Итого по отчету"][1]).replace(
                        ".", ","
                    ),
                    self.finreport_dict_beach["Депозит"][1],
                    self.finreport_dict_beach["Карты"][0],
                    self.finreport_dict_beach["Карты"][1],
                    f"=IFERROR(I{self.nex_line}/H{self.nex_line};0)",
                    self.finreport_dict_beach["Услуги"][0],
                    self.finreport_dict_beach["Услуги"][1],
                    f"=IFERROR(L{self.nex_line}/K{self.nex_line};0)",
                    self.finreport_dict_beach["Товары"][0],
                    self.finreport_dict_beach["Товары"][1],
                    f"=IFERROR(O{self.nex_line}/N{self.nex_line};0)",
                ]
            ],
            "ROWS",
        )

        # Задание форматы вывода строки
        ss.prepare_setCellsFormats(
            f"A{self.nex_line}:P{self.nex_line}",
            [
                [
                    {
                        "numberFormat": {"type": "DATE", "pattern": "dd.mm.yyyy"},
                        "horizontalAlignment": "LEFT",
                    },
                    {"numberFormat": {}},
                    {"numberFormat": {}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                    {"numberFormat": {}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                    {"numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00[$ ₽]"}},
                ]
            ],
        )
        # Цвет фона ячеек
        if self.nex_line % 2 != 0:
            ss.prepare_setCellsFormat(
                f"A{self.nex_line}:P{self.nex_line}",
                {"backgroundColor": functions.htmlColorToJSON("#fef8e3")},
                fields="userEnteredFormat.backgroundColor",
            )

        # Бордер
        for j in range(self.sheet6_width):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": self.nex_line - 1,
                            "endRowIndex": self.nex_line,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )

        # ------------------------------------------- Заполнение ИТОГО --------------------------------------
        logger.info("Заполнение строки ИТОГО на листе 2...")

        for i, line_table in enumerate(
            self.spreadsheet["sheets"][1]["data"][0]["rowData"]
        ):
            try:
                if line_table["values"][0]["formattedValue"] == "ИТОГО":
                    # Если строка переписывается - итого на 1 поз вниз, если новая - на 2 поз
                    height_table = i + self.reprint
                    break
                else:
                    height_table = 4
            except KeyError:
                pass

        ss.prepare_setValues(
            f"A{height_table}:P{height_table}",
            [
                [
                    "ИТОГО",
                    "",
                    f"=SUM(C3:C{height_table - 1})",
                    f"=SUM(D3:D{height_table - 1})",
                    f"=SUM(E3:E{height_table - 1})",
                    f"=SUM(F3:F{height_table - 1})",
                    f"=SUM(G3:G{height_table - 1})",
                    f"=SUM(H3:H{height_table - 1})",
                    f"=SUM(I3:I{height_table - 1})",
                    f"=IFERROR(ROUND(I{height_table}/H{height_table};2);0)",
                    f"=SUM(K3:K{height_table - 1})",
                    f"=SUM(L3:L{height_table - 1})",
                    f"=IFERROR(ROUND(L{height_table}/K{height_table};2);0)",
                    f"=SUM(N3:N{height_table - 1})",
                    f"=SUM(O3:O{height_table - 1})",
                    f"=IFERROR(ROUND(O{height_table}/N{height_table};2);0)",
                ]
            ],
            "ROWS",
        )
        ss.prepare_setValues(
            f"A{height_table + 1}:D{height_table + 1}",
            [
                [
                    "Выполнение плана (трафик)",
                    "",
                    f"=IFERROR('План'!L{self.sheet2_line};0)",
                    f"=IFERROR(ROUND(D{height_table}/C{height_table + 1};2);0)",
                ]
            ],
            "ROWS",
        )
        ss.prepare_setValues(
            f"A{height_table + 2}:D{height_table + 2}",
            [
                [
                    "Выполнение плана (доход)",
                    "",
                    f"=IFERROR('План'!M{self.sheet2_line};0)",
                    f"=IFERROR(ROUND(F{height_table}/C{height_table + 2};2);0)",
                ]
            ],
            "ROWS",
        )

        # Задание формата вывода строки
        ss.prepare_setCellsFormats(
            f"A{height_table}:P{height_table}",
            [
                [
                    {"textFormat": {"bold": True}},
                    {"textFormat": {"bold": True}},
                    {"textFormat": {"bold": True}},
                    {"textFormat": {"bold": True}},
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {"horizontalAlignment": "RIGHT", "textFormat": {"bold": True}},
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                    {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                        "horizontalAlignment": "RIGHT",
                        "textFormat": {"bold": True},
                    },
                ]
            ],
        )
        ss.prepare_setCellsFormats(
            f"A{height_table + 1}:D{height_table + 1}",
            [
                [
                    {"textFormat": {"bold": True}},
                    {"textFormat": {"bold": True}},
                    {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "RIGHT",
                        "numberFormat": {},
                    },
                    {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "RIGHT",
                        "numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00%"},
                    },
                ]
            ],
        )
        ss.prepare_setCellsFormats(
            f"A{height_table + 2}:D{height_table + 2}",
            [
                [
                    {"textFormat": {"bold": True}},
                    {"textFormat": {"bold": True}},
                    {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "RIGHT",
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "#,##0.00[$ ₽]",
                        },
                    },
                    {
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "RIGHT",
                        "numberFormat": {"type": "CURRENCY", "pattern": "#,##0.00%"},
                    },
                ]
            ],
        )

        # Цвет фона ячеек
        ss.prepare_setCellsFormat(
            f"A{height_table}:P{height_table}",
            {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
            fields="userEnteredFormat.backgroundColor",
        )
        ss.prepare_setCellsFormat(
            f"A{height_table + 1}:D{height_table + 1}",
            {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
            fields="userEnteredFormat.backgroundColor",
        )
        ss.prepare_setCellsFormat(
            f"A{height_table + 2}:D{height_table + 2}",
            {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
            fields="userEnteredFormat.backgroundColor",
        )

        # Бордер
        for j in range(self.sheet6_width):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table - 1,
                            "endRowIndex": height_table,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
        for j in range(4):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table,
                            "endRowIndex": height_table + 1,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table,
                            "endRowIndex": height_table + 1,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table,
                            "endRowIndex": height_table + 1,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table,
                            "endRowIndex": height_table + 1,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
        for j in range(4):
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table + 1,
                            "endRowIndex": height_table + 2,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "top": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table + 1,
                            "endRowIndex": height_table + 2,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "right": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table + 1,
                            "endRowIndex": height_table + 2,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "left": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
            ss.requests.append(
                {
                    "updateBorders": {
                        "range": {
                            "sheetId": ss.sheetId,
                            "startRowIndex": height_table + 1,
                            "endRowIndex": height_table + 2,
                            "startColumnIndex": j,
                            "endColumnIndex": j + 1,
                        },
                        "bottom": {
                            "style": "SOLID",
                            "width": 1,
                            "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0},
                        },
                    }
                }
            )
        ss.runPrepared()

    def sms_report(self, date_from) -> str:
        """Составляет текстовую версию финансового отчета."""

        logger.info("Составление SMS-отчета...")
        resporse = "Отчет по аквапарку за "

        if self.finreport_dict["Дата"][0] == self.finreport_dict["Дата"][1] - timedelta(
            1
        ):
            resporse += (
                f'{datetime.strftime(self.finreport_dict["Дата"][0], "%d.%m.%Y")}:\n'
            )

        else:
            resporse += (
                f'{datetime.strftime(self.finreport_dict["Дата"][0], "%d.%m.%Y")} '
                f'- {datetime.strftime(self.finreport_dict["Дата"][1] - timedelta(1), "%d.%m.%Y")}:\n'
            )

        def get_sum(field_name: str) -> float:
            return self.finreport_dict.get(field_name, [0, 0])[1]

        bars_sum = get_sum("ИТОГО")
        smile = get_sum("Смайл")
        bonuses = get_sum("MaxBonus")
        other = get_sum("Прочее") + get_sum("Сопутствующие товары")
        total_sum = bars_sum - bonuses + smile

        if self.finreport_dict["ИТОГО"][1]:
            resporse += f'Люди - {self.finreport_dict["Кол-во проходов"][0]};\n'
            resporse += f"По аквапарку - {get_sum('Билеты аквапарка') + get_sum('Билеты аквапарка КОРП'):.2f} ₽;\n"
            resporse += f"По общепиту - {(get_sum('Общепит') + smile):.2f} ₽;\n"

            resporse += f"Прочее - {other:.2f} ₽;\n"
            resporse += f"Общая по БАРСу - {bars_sum:.2f} ₽;\n"
            resporse += f"Бонусы - {bonuses:.2f} ₽;\n"
            resporse += f"ONLINE продажи - {get_sum('Online Продажи'):.2f} ₽;\n"

        if not re.search(r"По общепиту", resporse) and smile:
            resporse += f"По общепиту - {smile:.2f} ₽;\n"
            resporse += f"Общая по БАРСу - {bars_sum:.2f} ₽;\n"

        resporse += f"Общая ИТОГО - {total_sum:.2f} ₽;\n\n"

        if self.itog_report_org2["Итого по отчету"][1]:
            try:
                resporse += f'Люди (пляж) - {self.itog_report_org2["Летняя зона | БЕЗЛИМИТ | 1 проход"][0]};\n'
            except KeyError:
                pass
            resporse += f'Итого по пляжу - {self.itog_report_org2["Итого по отчету"][1]:.2f} ₽;\n'

        resporse += "Без ЧП."

        with open(
            f'reports/{date_from.strftime("%Y.%m.%d")}_sms.txt', "w", encoding="utf-8"
        ) as f:
            f.write(resporse)
        return resporse

    def save_organisation_total(self, itog_report, date_from):
        """
        Сохраняет Итоговый отчет в Excel
        """
        organisation_total = {}
        for key in itog_report:
            organisation_total[itog_report[key][3]] = {}
        for key in itog_report:
            organisation_total[itog_report[key][3]][itog_report[key][2]] = []
        for key in itog_report:
            organisation_total[itog_report[key][3]][itog_report[key][2]].append(
                (key, itog_report[key][0], itog_report[key][1])
            )

        # определяем стили
        h1 = Font(
            name="Times New Roman",
            size=18,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        h2 = Font(
            name="Times New Roman",
            size=14,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        h3 = Font(
            name="Times New Roman",
            size=11,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font = Font(
            name="Times New Roman",
            size=9,
            bold=False,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font_bold = Font(
            name="Times New Roman",
            size=9,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )

        fill = PatternFill(fill_type="solid", start_color="c1c1c1", end_color="c2c2c2")
        align_top = Alignment(
            horizontal="general",
            vertical="top",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )
        align_bottom = Alignment(
            horizontal="general",
            vertical="bottom",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )

        border = Border(
            left=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
            top=Side(border_style="thin", color="FF000000"),
            bottom=Side(border_style="thin", color="FF000000"),
            diagonal=Side(border_style="thin", color="FF000000"),
            diagonal_direction=0,
            outline=Side(border_style="thin", color="FF000000"),
            vertical=Side(border_style="thin", color="FF000000"),
            horizontal=Side(border_style="thin", color="FF000000"),
        )
        border_top_bottom = Border(
            bottom=Side(border_style="thin", color="FF000000"),
            top=Side(border_style="thin", color="FF000000"),
        )
        border_right = Border(right=Side(border_style="thin", color="FF000000"))
        border_left = Border(left=Side(border_style="thin", color="FF000000"))
        border_top = Border(top=Side(border_style="thin", color="FF000000"))
        border_left_top = Border(
            top=Side(border_style="thin", color="FF000000"),
            left=Side(border_style="thin", color="FF000000"),
        )
        border_right_top = Border(
            top=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
        )
        align_left = Alignment(
            horizontal="left",
            vertical="bottom",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )
        column = ["", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Итоговый отчет"
        # шрифты
        ws["C1"].font = h1
        # выравнивание
        ws["C1"].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 8
        ws.column_dimensions["B"].width = 1 / 7 * 8
        ws.column_dimensions["C"].width = 1 / 7 * 80
        ws.column_dimensions["D"].width = 1 / 7 * 8
        ws.column_dimensions["E"].width = 1 / 7 * 88
        ws.column_dimensions["F"].width = 1 / 7 * 8
        ws.column_dimensions["G"].width = 1 / 7 * 24
        ws.column_dimensions["H"].width = 1 / 7 * 8
        ws.column_dimensions["I"].width = 1 / 7 * 80
        ws.column_dimensions["J"].width = 1 / 7 * 8
        ws.column_dimensions["K"].width = 1 / 7 * 144
        ws.column_dimensions["L"].width = 1 / 7 * 144
        ws.column_dimensions["M"].width = 1 / 7 * 8

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[3] + next_row()] = "Итоговый отчет"
        ws.merge_cells(
            start_row=self.row, start_column=3, end_row=self.row, end_column=12
        )
        ws[column[1] + next_row()] = ""
        ws[column[3] + next_row()] = organisation_total["Организация"]["Организация"][
            0
        ][0]
        ws.merge_cells(
            start_row=self.row, start_column=3, end_row=self.row, end_column=12
        )
        ws[column[3] + self.row].font = font
        ws[column[3] + self.row].alignment = align_top
        ws[column[1] + next_row()] = ""

        ws[column[3] + next_row()] = "За период с:"
        ws[column[3] + self.row].font = font
        ws[column[3] + self.row].alignment = align_top
        ws[column[5] + self.row] = itog_report["Дата"][0].strftime("%d.%m.%Y")
        ws[column[5] + self.row].font = font_bold
        ws[column[5] + self.row].alignment = align_top
        ws[column[7] + self.row] = "По:"
        ws[column[7] + self.row].font = font
        ws[column[7] + self.row].alignment = align_top
        ws[column[9] + self.row] = (itog_report["Дата"][1] - timedelta(1)).strftime(
            "%d.%m.%Y"
        )
        ws[column[9] + self.row].font = font_bold
        ws[column[9] + self.row].alignment = align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(
                start_row=self.row, start_column=2, end_row=self.row, end_column=9
            )
            ws.merge_cells(
                start_row=self.row, start_column=10, end_row=self.row, end_column=11
            )
            ws.merge_cells(
                start_row=self.row, start_column=12, end_row=self.row, end_column=13
            )
            ws[column[2] + self.row].font = font
            ws[column[10] + self.row].font = font
            ws[column[12] + self.row].font = font
            ws[column[2] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[12] + self.row].alignment = align_top
            b = 2
            while b <= 13:
                ws[column[b] + self.row].border = border
                b += 1

        def merge_table_h3():
            ws.merge_cells(
                start_row=self.row, start_column=2, end_row=self.row, end_column=9
            )
            ws.merge_cells(
                start_row=self.row, start_column=10, end_row=self.row, end_column=11
            )
            ws.merge_cells(
                start_row=self.row, start_column=12, end_row=self.row, end_column=13
            )
            ws[column[2] + self.row].font = h3
            ws[column[10] + self.row].font = h3
            ws[column[12] + self.row].font = h3
            ws[column[2] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[12] + self.row].alignment = align_top
            ws[column[2] + self.row].border = border_left
            ws[column[13] + self.row].border = border_right

        def merge_table_h2():
            ws.merge_cells(
                start_row=self.row, start_column=2, end_row=self.row, end_column=9
            )
            ws.merge_cells(
                start_row=self.row, start_column=10, end_row=self.row, end_column=11
            )
            ws.merge_cells(
                start_row=self.row, start_column=12, end_row=self.row, end_column=13
            )
            ws[column[2] + self.row].font = h2
            ws[column[10] + self.row].font = h2
            ws[column[12] + self.row].font = h2
            ws[column[2] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[12] + self.row].alignment = align_top
            ws[column[2] + self.row].border = border_left
            ws[column[13] + self.row].border = border_right

        def merge_width_h2():
            ws.merge_cells(
                start_row=self.row, start_column=2, end_row=self.row, end_column=13
            )
            ws[column[2] + self.row].font = h2
            ws[column[2] + self.row].alignment = align_top
            b = 2
            while b <= 13:
                if b == 2:
                    ws[column[b] + self.row].border = border_left_top
                elif b == 13:
                    ws[column[b] + self.row].border = border_right_top
                else:
                    ws[column[b] + self.row].border = border_top
                b += 1

        def merge_width_h3():
            ws.merge_cells(
                start_row=self.row, start_column=2, end_row=self.row, end_column=13
            )
            ws[column[2] + self.row].font = h3
            ws[column[2] + self.row].alignment = align_top
            b = 2
            while b <= 13:
                if b == 2:
                    ws[column[b] + self.row].border = border_left
                elif b == 13:
                    ws[column[b] + self.row].border = border_right
                b += 1

        ws[column[2] + next_row()] = "Название"
        ws[column[10] + self.row] = "Количество"
        ws[column[12] + self.row] = "Сумма"
        merge_table()
        ws[column[2] + self.row].font = h3
        ws[column[10] + self.row].font = h3
        ws[column[12] + self.row].font = h3
        ws[column[2] + self.row].alignment = align_top
        ws[column[10] + self.row].alignment = align_top
        ws[column[12] + self.row].alignment = align_top

        groups = [
            "Депозит",
            "Карты",
            "Услуги",
            "Товары",
            "Платные зоны",
        ]
        all_count = 0
        all_sum = 0
        try:
            for gr in groups:
                ws[column[2] + next_row()] = gr
                merge_width_h2()
                group_count = 0
                group_sum = 0
                organisation_total_groups = organisation_total.get(gr)
                if not organisation_total_groups:
                    continue
                for group in organisation_total_groups:
                    ws[column[2] + next_row()] = group
                    merge_width_h3()
                    service_count = 0
                    service_sum = 0
                    servises = organisation_total_groups.get(group, [])
                    if not servises:
                        continue
                    for service in servises:
                        try:
                            service_count += service[1]
                            service_sum += service[2]
                        except TypeError:
                            pass
                        ws[column[2] + next_row()] = service[0]
                        ws[column[10] + self.row] = service[1]
                        ws[column[12] + self.row] = service[2]
                        ws[column[12] + self.row].number_format = "#,##0.00 ₽"
                        merge_table()
                    ws[column[10] + next_row()] = service_count
                    ws[column[12] + self.row] = service_sum
                    ws[column[12] + self.row].number_format = "#,##0.00 ₽"
                    merge_table_h3()
                    group_count += service_count
                    group_sum += service_sum
                ws[column[10] + next_row()] = group_count
                ws[column[12] + self.row] = group_sum
                ws[column[12] + self.row].number_format = "#,##0.00 ₽"
                merge_table_h2()
                all_count += group_count
                all_sum += group_sum
                group_count = 0
                group_sum = 0
        except KeyError:
            pass

        bars_total_sum = organisation_total["Итого по отчету"][""][0]
        if all_sum == bars_total_sum[2]:
            ws[column[2] + next_row()] = bars_total_sum[0]
            ws[column[10] + self.row] = bars_total_sum[1]
            ws[column[12] + self.row] = bars_total_sum[2]
            self.total_report_sum = all_sum
        else:
            error_code = "Ошибка: Итоговые суммы не совпадают."
            error_message = (
                f'"Итого по отчету" из Барса ({bars_total_sum[2]})'
                f" не совпадает с итоговой суммой по формируемым строкам ({all_sum})."
            )
            logger.error(f"{error_code} {error_message}")
            return None

        ws[column[12] + self.row].number_format = "#,##0.00 ₽"
        merge_table_h2()
        ws[column[2] + self.row].alignment = align_bottom
        ws[column[10] + self.row].alignment = align_bottom
        ws[column[12] + self.row].alignment = align_bottom
        b = 2
        while b <= 13:
            ws[column[b] + self.row].border = border_top_bottom
            b += 1
        end_line = int(self.row)

        # раскрвшивание фона для заголовков
        i = 2
        while i <= 13:
            ws[column[i] + "6"].fill = fill
            i += 1

        # обводка
        # ws['A3'].border = border

        # вручную устанавливаем высоту первой строки
        rd = ws.row_dimensions[1]
        rd.height = 21.75

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1

        # Высота строк
        ws.row_dimensions[2].height = 5.25
        ws.row_dimensions[4].height = 6.75
        ws.row_dimensions[end_line].height = 30.75

        # выравнивание столбца
        for cellObj in ws["A2:A5"]:
            for cell in cellObj:
                ws[cell.coordinate].alignment = align_left

        # перетягивание ячеек
        # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
        # dims = {}
        # for row in ws.rows:
        #     for cell in row:
        #         if cell.value:
        #             dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        # for col, value in dims.items():
        #     # value * коэфициент
        #     ws.column_dimensions[col].width = value * 1.5

        # сохранение файла в текущую директорию
        if itog_report["Дата"][0] == itog_report["Дата"][1] - timedelta(1):
            date_ = datetime.strftime(itog_report["Дата"][0], "%Y-%m-%d")
        else:
            date_ = (
                f'{datetime.strftime(itog_report["Дата"][0], "%Y-%m-%d")} - '
                f'{datetime.strftime(itog_report["Дата"][1] - timedelta(1), "%Y-%m-%d")}'
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f' Итоговый отчет по {organisation_total["Организация"]["Организация"][0][0]} '
            + ".xlsx"
        )
        logger.info(
            f"Сохранение Итогового отчета "
            f'по {organisation_total["Организация"]["Организация"][0][0]} в {path}'
        )
        path = self._yandex_repo.create_path(path, date_from)
        self._yandex_repo.save_file(path, wb)
        return path

    def save_cashdesk_report(self, cashdesk_report, date_from):
        """
        Сохраняет Суммовой отчет в Excel
        """
        # определяем стили
        h1 = Font(
            name="Times New Roman",
            size=18,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        h3 = Font(
            name="Times New Roman",
            size=10,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font = Font(
            name="Times New Roman",
            size=9,
            bold=False,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font_bold = Font(
            name="Times New Roman",
            size=9,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font_red = Font(
            name="Times New Roman",
            size=9,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FFFF0000",
        )

        fill = PatternFill(fill_type="solid", start_color="c1c1c1", end_color="c2c2c2")
        align_top = Alignment(
            horizontal="general",
            vertical="top",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )
        border = Border(
            left=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
            top=Side(border_style="thin", color="FF000000"),
            bottom=Side(border_style="thin", color="FF000000"),
            diagonal=Side(border_style="thin", color="FF000000"),
            diagonal_direction=0,
            outline=Side(border_style="thin", color="FF000000"),
            vertical=Side(border_style="thin", color="FF000000"),
            horizontal=Side(border_style="thin", color="FF000000"),
        )
        border_right = Border(right=Side(border_style="thin", color="FF000000"))
        border_left = Border(left=Side(border_style="thin", color="FF000000"))
        align_left = Alignment(
            horizontal="left",
            vertical="bottom",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )
        column = [
            "",
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
        ]

        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Суммовой отчет по чековой ленте"
        # шрифты
        ws["A1"].font = h1
        # выравнивание
        ws["A1"].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 124
        ws.column_dimensions["B"].width = 1 / 7 * 88
        ws.column_dimensions["C"].width = 1 / 7 * 28
        ws.column_dimensions["D"].width = 1 / 7 * 24
        ws.column_dimensions["E"].width = 1 / 7 * 32
        ws.column_dimensions["F"].width = 1 / 7 * 1
        ws.column_dimensions["G"].width = 1 / 7 * 79
        ws.column_dimensions["H"].width = 1 / 7 * 3
        ws.column_dimensions["I"].width = 1 / 7 * 5
        ws.column_dimensions["J"].width = 1 / 7 * 96
        ws.column_dimensions["K"].width = 1 / 7 * 88
        ws.column_dimensions["L"].width = 1 / 7 * 8
        ws.column_dimensions["M"].width = 1 / 7 * 80
        ws.column_dimensions["N"].width = 1 / 7 * 96

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = "Суммовой отчет по чековой ленте"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        # шрифты
        ws[column[1] + self.row].font = h1
        # выравнивание
        ws[column[1] + self.row].alignment = align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f'{cashdesk_report["Организация"][0][0]}'
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top

        ws[column[1] + next_row()] = "За период с:"
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top
        ws[column[2] + self.row] = cashdesk_report["Дата"][0][0].strftime("%d.%m.%Y")
        ws.merge_cells(
            start_row=self.row, start_column=2, end_row=self.row, end_column=3
        )
        ws[column[2] + self.row].font = font_bold
        ws[column[2] + self.row].alignment = align_top
        ws[column[4] + self.row] = "по"
        ws[column[4] + self.row].font = font
        ws[column[4] + self.row].alignment = align_top
        ws[column[5] + self.row] = (
            cashdesk_report["Дата"][0][1] - timedelta(1)
        ).strftime("%d.%m.%Y")
        ws.merge_cells(
            start_row=self.row, start_column=5, end_row=self.row, end_column=7
        )
        ws[column[5] + self.row].font = font_bold
        ws[column[5] + self.row].alignment = align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(
                start_row=self.row, start_column=1, end_row=self.row, end_column=2
            )
            ws.merge_cells(
                start_row=self.row, start_column=3, end_row=self.row, end_column=6
            )
            ws.merge_cells(
                start_row=self.row, start_column=7, end_row=self.row, end_column=9
            )
            ws.merge_cells(
                start_row=self.row, start_column=11, end_row=self.row, end_column=12
            )
            ws[column[1] + self.row].font = font
            ws[column[3] + self.row].font = font
            ws[column[7] + self.row].font = font
            ws[column[10] + self.row].font = font
            ws[column[11] + self.row].font = font
            ws[column[13] + self.row].font = font
            ws[column[14] + self.row].font = font
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            ws[column[7] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[11] + self.row].alignment = align_top
            ws[column[13] + self.row].alignment = align_top
            ws[column[14] + self.row].alignment = align_top
            ws[column[3] + self.row].number_format = "#,##0.00 ₽"
            ws[column[7] + self.row].number_format = "#,##0.00 ₽"
            ws[column[10] + self.row].number_format = "#,##0.00 ₽"
            ws[column[11] + self.row].number_format = "#,##0.00 ₽"
            ws[column[13] + self.row].number_format = "#,##0.00 ₽"
            ws[column[14] + self.row].number_format = "#,##0.00 ₽"
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        def merge_table_h3():
            ws.merge_cells(
                start_row=self.row, start_column=1, end_row=self.row, end_column=2
            )
            ws.merge_cells(
                start_row=self.row, start_column=3, end_row=self.row, end_column=6
            )
            ws.merge_cells(
                start_row=self.row, start_column=7, end_row=self.row, end_column=9
            )
            ws.merge_cells(
                start_row=self.row, start_column=11, end_row=self.row, end_column=12
            )
            ws[column[1] + self.row].font = h3
            ws[column[3] + self.row].font = h3
            ws[column[7] + self.row].font = h3
            ws[column[10] + self.row].font = h3
            ws[column[11] + self.row].font = h3
            ws[column[13] + self.row].font = h3
            ws[column[14] + self.row].font = h3
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            ws[column[7] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[11] + self.row].alignment = align_top
            ws[column[13] + self.row].alignment = align_top
            ws[column[14] + self.row].alignment = align_top
            ws[column[3] + self.row].number_format = "#,##0.00 ₽"
            ws[column[7] + self.row].number_format = "#,##0.00 ₽"
            ws[column[10] + self.row].number_format = "#,##0.00 ₽"
            ws[column[11] + self.row].number_format = "#,##0.00 ₽"
            ws[column[13] + self.row].number_format = "#,##0.00 ₽"
            ws[column[14] + self.row].number_format = "#,##0.00 ₽"
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        def merge_table_red():
            ws.merge_cells(
                start_row=self.row, start_column=1, end_row=self.row, end_column=2
            )
            ws.merge_cells(
                start_row=self.row, start_column=3, end_row=self.row, end_column=6
            )
            ws.merge_cells(
                start_row=self.row, start_column=7, end_row=self.row, end_column=9
            )
            ws.merge_cells(
                start_row=self.row, start_column=11, end_row=self.row, end_column=12
            )
            ws[column[1] + self.row].font = font_red
            ws[column[3] + self.row].font = font_red
            ws[column[7] + self.row].font = font_red
            ws[column[10] + self.row].font = font_red
            ws[column[11] + self.row].font = font_red
            ws[column[13] + self.row].font = font_red
            ws[column[14] + self.row].font = font_red
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            ws[column[7] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[11] + self.row].alignment = align_top
            ws[column[13] + self.row].alignment = align_top
            ws[column[14] + self.row].alignment = align_top
            ws[column[3] + self.row].number_format = "#,##0.00 ₽"
            ws[column[7] + self.row].number_format = "#,##0.00 ₽"
            ws[column[10] + self.row].number_format = "#,##0.00 ₽"
            ws[column[11] + self.row].number_format = "#,##0.00 ₽"
            ws[column[13] + self.row].number_format = "#,##0.00 ₽"
            ws[column[14] + self.row].number_format = "#,##0.00 ₽"
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        def merge_width_red():
            ws.merge_cells(
                start_row=self.row,
                start_column=1,
                end_row=self.row,
                end_column=len(column) - 1,
            )
            ws[column[1] + self.row].font = font_red
            ws[column[1] + self.row].alignment = align_top
            b = 1
            while b < len(column):
                if b == 1:
                    ws[column[b] + self.row].border = border_left
                elif b == len(column) - 1:
                    ws[column[b] + self.row].border = border_right
                else:
                    ws[column[b] + self.row].border = border
                b += 1

        ws[column[1] + next_row()] = "Касса №"
        ws[column[3] + self.row] = "Сумма"
        ws[column[7] + self.row] = "Наличными"
        ws[column[10] + self.row] = "Безналичными"
        ws[column[11] + self.row] = "Со счета"
        ws[column[13] + self.row] = "Бонусами"
        ws[column[14] + self.row] = "MaxBonux"
        merge_table_h3()
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = fill
            b += 1

        for typpe in cashdesk_report:
            if typpe != "Дата" and typpe != "Организация":
                if typpe != "Итого":
                    ws[column[1] + next_row()] = typpe
                    merge_width_red()
                for line in cashdesk_report[typpe]:
                    ws[column[1] + next_row()] = line[0]
                    ws[column[3] + self.row] = line[1]
                    ws[column[7] + self.row] = line[2]
                    ws[column[10] + self.row] = line[3]
                    ws[column[11] + self.row] = line[4]
                    ws[column[13] + self.row] = line[5]
                    ws[column[14] + self.row] = line[6] - line[3]
                    if line[0] == "Итого":
                        merge_table_red()
                    elif line[0] == "Итого по отчету":
                        merge_table_h3()
                    else:
                        merge_table()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if cashdesk_report["Дата"][0][0] == cashdesk_report["Дата"][0][1] - timedelta(
            1
        ):
            date_ = datetime.strftime(cashdesk_report["Дата"][0][0], "%Y-%m-%d")
        else:
            date_ = (
                f'{datetime.strftime(cashdesk_report["Дата"][0][0], "%Y-%m-%d")} - '
                f'{datetime.strftime(cashdesk_report["Дата"][0][1] - timedelta(1), "%Y-%m-%d")}'
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f' Суммовой отчет по {cashdesk_report["Организация"][0][0]}'
            + ".xlsx"
        )
        logger.info(
            f"Сохранение Суммового отчета "
            f'по {cashdesk_report["Организация"][0][0]} в {path}'
        )
        path = self._yandex_repo.create_path(path, date_from)
        self._yandex_repo.save_file(path, wb)
        return path

    def save_client_count_totals(self, client_count_totals_org, date_from):
        """
        Сохраняет отчет по количеству клиентов за день в Excel
        """
        # определяем стили
        h1 = Font(
            name="Times New Roman",
            size=18,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font = Font(
            name="Times New Roman",
            size=9,
            bold=False,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        font_bold = Font(
            name="Times New Roman",
            size=9,
            bold=True,
            italic=False,
            vertAlign=None,
            underline="none",
            strike=False,
            color="FF000000",
        )
        fill = PatternFill(fill_type="solid", start_color="c1c1c1", end_color="c2c2c2")
        align_top = Alignment(
            horizontal="general",
            vertical="top",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )
        border = Border(
            left=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
            top=Side(border_style="thin", color="FF000000"),
            bottom=Side(border_style="thin", color="FF000000"),
            diagonal=Side(border_style="thin", color="FF000000"),
            diagonal_direction=0,
            outline=Side(border_style="thin", color="FF000000"),
            vertical=Side(border_style="thin", color="FF000000"),
            horizontal=Side(border_style="thin", color="FF000000"),
        )
        align_left = Alignment(
            horizontal="left",
            vertical="bottom",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )
        column = ["", "A", "B", "C", "D", "E"]

        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Количество человек за день"
        # шрифты
        ws["A1"].font = h1
        # выравнивание
        ws["A1"].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 124
        ws.column_dimensions["B"].width = 1 / 7 * 21
        ws.column_dimensions["C"].width = 1 / 7 * 95
        ws.column_dimensions["D"].width = 1 / 7 * 24
        ws.column_dimensions["E"].width = 1 / 7 * 80

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = "Количество человек за день"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        # шрифты
        ws[column[1] + self.row].font = h1
        # выравнивание
        ws[column[1] + self.row].alignment = align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f"{client_count_totals_org[0][0]}"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top

        ws[column[1] + next_row()] = "За период с:"
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top
        ws[column[2] + self.row] = client_count_totals_org[1][0].strftime("%d.%m.%Y")
        ws.merge_cells(
            start_row=self.row, start_column=2, end_row=self.row, end_column=3
        )
        ws[column[2] + self.row].font = font_bold
        ws[column[2] + self.row].alignment = align_top
        ws[column[4] + self.row] = "по"
        ws[column[4] + self.row].font = font
        ws[column[4] + self.row].alignment = align_top
        ws[column[5] + self.row] = (client_count_totals_org[-2][0]).strftime("%d.%m.%Y")
        ws.merge_cells(
            start_row=self.row, start_column=5, end_row=self.row, end_column=7
        )
        ws[column[5] + self.row].font = font_bold
        ws[column[5] + self.row].alignment = align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(
                start_row=self.row, start_column=1, end_row=self.row, end_column=2
            )
            ws.merge_cells(
                start_row=self.row, start_column=3, end_row=self.row, end_column=5
            )
            ws[column[1] + self.row].font = font
            ws[column[3] + self.row].font = font
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        def merge_table_bold():
            ws.merge_cells(
                start_row=self.row, start_column=1, end_row=self.row, end_column=2
            )
            ws.merge_cells(
                start_row=self.row, start_column=3, end_row=self.row, end_column=5
            )
            ws[column[1] + self.row].font = font_bold
            ws[column[3] + self.row].font = font_bold
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        ws[column[1] + next_row()] = "Дата"
        ws[column[3] + self.row] = "Количество клиентов"
        merge_table_bold()
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = fill
            b += 1

        for line in client_count_totals_org:
            try:
                ws[column[1] + next_row()] = line[0].strftime("%d.%m.%Y")
                ws[column[3] + self.row] = line[1]
                merge_table()
            except AttributeError:
                pass

        ws[column[1] + next_row()] = "Итого"
        ws[column[3] + self.row] = client_count_totals_org[-1][1]
        merge_table_bold()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if client_count_totals_org[0][1]:
            date_ = datetime.strftime(client_count_totals_org[1][0], "%Y-%m")
        else:
            date_ = (
                f'{datetime.strftime(client_count_totals_org[1][0], "%Y-%m-%d")} - '
                f'{datetime.strftime(client_count_totals_org[-2][0], "%Y-%m-%d")}'
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f" Количество клиентов за день по {client_count_totals_org[0][0]}"
            + ".xlsx"
        )
        logger.info(
            f"Сохранение отчета по количеству клиентов "
            f"по {client_count_totals_org[0][0]} в {path}"
        )
        path = self._yandex_repo.create_path(path, date_from)
        self._yandex_repo.save_file(path, wb)
        return path

    async def save_reports(self, date_from):
        """
        Функция управления
        """
        self.fin_report()
        self.agent_report()
        # agentreport_xls
        self.path_list.append(
            self.export_agent_report(self.agentreport_dict, date_from)
        )
        # finreport_google
        self.fin_report_lastyear()
        self.fin_report_beach()

        self.finreport_dict_month = None
        if self.itog_report_month:
            self.finreport_dict_month = functions.create_month_finance_report(
                itog_report_month=self.itog_report_month,
                itogreport_group_dict=self.itogreport_group_dict,
                orgs_dict=self.orgs_dict,
                smile_report_month=self.smile_report_month,
            )
            self.agentreport_dict_month = functions.create_month_agent_report(
                month_total_report=self.itog_report_month,
                agent_dict=self.agent_dict,
            )

        credentials = ServiceAccountCredentials.from_json_keyfile_dict(
            settings.google_api_settings.google_service_account_config,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ],
        )
        httpAuth = credentials.authorize(httplib2.Http())
        try:
            logger.info("Попытка авторизации с Google-документами ...")
            googleservice = apiclient.discovery.build(
                "sheets", "v4", http=httpAuth, cache_discovery=False
            )

        except IndexError as e:
            error_message = f"Ошибка {repr(e)}"
            logger.error(error_message)
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail=error_message,
            )

        await self.export_to_google_sheet(date_from, httpAuth, googleservice)
        # finreport_telegram:
        self.sms_report_list.append(self.sms_report(date_from))
        # check_itogreport_xls:
        if self.itog_report_org1["Итого по отчету"][1]:
            self.path_list.append(
                self.save_organisation_total(self.itog_report_org1, date_from)
            )
        if self.itog_report_org2["Итого по отчету"][1]:
            self.path_list.append(
                self.save_organisation_total(self.itog_report_org2, date_from)
            )
        if self.itog_report_org3["Итого по отчету"][1]:
            self.path_list.append(
                self.save_organisation_total(self.itog_report_org3, date_from)
            )
        if self.itog_report_org4["Итого по отчету"][1]:
            self.path_list.append(
                self.save_organisation_total(self.itog_report_org4, date_from)
            )
        if self.itog_report_org5["Итого по отчету"][1]:
            self.path_list.append(
                self.save_organisation_total(self.itog_report_org5, date_from)
            )
        # check_cashreport_xls:
        if self.cashdesk_report_org1["Итого"][0][1]:
            self.path_list.append(
                self.save_cashdesk_report(self.cashdesk_report_org1, date_from)
            )
        if self.cashdesk_report_org2["Итого"][0][1]:
            self.path_list.append(
                self.save_cashdesk_report(self.cashdesk_report_org2, date_from)
            )
        # check_client_count_total_xls:
        if self.client_count_totals_org1[-1][1]:
            self.path_list.append(
                self.save_client_count_totals(self.client_count_totals_org1, date_from)
            )
        if self.client_count_totals_org2[-1][1]:
            self.path_list.append(
                self.save_client_count_totals(self.client_count_totals_org2, date_from)
            )

    async def load_report(self, date_from, date_to):
        """Выполнить отчеты"""

        self.itog_report_org1 = None
        self.itog_report_org2 = None
        self.itog_report_org3 = None
        self.itog_report_org4 = None
        self.itog_report_org5 = None

        self.click_select_org()

        self.report_bitrix = (0, 0)
        self.report_bitrix_lastyear = (0, 0)
        self.smile_report = self._rk_service.get_smile_report(
            date_from=date_from,
            date_to=date_to,
        )
        self.smile_report_lastyear = self._rk_service.get_smile_report(
            date_from=date_from - relativedelta(years=1),
            date_to=date_to - relativedelta(years=1),
        )

        if self.org1:
            self.bars_srv.set_database(settings.mssql_database1)
            with self.bars_srv as connect:
                self.itog_report_org1 = functions.get_total_report(
                    connect=connect,
                    org=self.org1[0],
                    org_name=self.org1[1],
                    date_from=date_from,
                    date_to=date_to,
                )
                self.itog_report_org1_lastyear = functions.get_total_report(
                    connect=connect,
                    org=self.org1[0],
                    org_name=self.org1[1],
                    date_from=date_from - relativedelta(years=1),
                    date_to=date_to - relativedelta(years=1),
                )
                self.itog_report_org3 = functions.get_total_report(
                    connect=connect,
                    org=self.org3[0],
                    org_name=self.org3[1],
                    date_from=date_from,
                    date_to=date_to,
                )
                self.itog_report_org3_lastyear = functions.get_total_report(
                    connect=connect,
                    org=self.org3[0],
                    org_name=self.org3[1],
                    date_from=date_from - relativedelta(years=1),
                    date_to=date_to - relativedelta(years=1),
                )
                self.itog_report_org4 = functions.get_total_report(
                    connect=connect,
                    org=self.org4[0],
                    org_name=self.org4[1],
                    date_from=date_from,
                    date_to=date_to,
                )
                self.itog_report_org4_lastyear = functions.get_total_report(
                    connect=connect,
                    org=self.org4[0],
                    org_name=self.org4[1],
                    date_from=date_from - relativedelta(years=1),
                    date_to=date_to - relativedelta(years=1),
                )
                self.itog_report_org5 = functions.get_total_report(
                    connect=connect,
                    org=self.org5[0],
                    org_name=self.org5[1],
                    date_from=date_from,
                    date_to=date_to,
                )
                self.itog_report_org5_lastyear = functions.get_total_report(
                    connect=connect,
                    org=self.org5[0],
                    org_name=self.org5[1],
                    date_from=date_from - relativedelta(years=1),
                    date_to=date_to - relativedelta(years=1),
                )

            self.itog_report_month = None
            if int((date_to - timedelta(1)).strftime("%y%m")) < int(
                date_to.strftime("%y%m")
            ):
                self._bars_service.choose_db(settings.mssql_database1)
                organizations = self._bars_service.get_organisations()
                itog_report_month = {}
                for organization in organizations:
                    self.bars_srv.set_database(settings.mssql_database1)
                    with self.bars_srv as connect:
                        itog_report_month_for_org = functions.get_total_report(
                            connect=connect,
                            org=organization.super_account_id,
                            org_name=organization.descr,
                            date_from=datetime.strptime(
                                "01" + (date_to - timedelta(1)).strftime("%m%y"),
                                "%d%m%y",
                            ),
                            date_to=date_to,
                        )
                    itog_report_month = functions.concatenate_itog_reports(
                        itog_report_month, itog_report_month_for_org
                    )

                self.itog_report_month = itog_report_month
                self.smile_report_month = self._rk_service.get_smile_report(
                    date_from=datetime.strptime(
                        "01" + (date_to - timedelta(1)).strftime("%m%y"), "%d%m%y"
                    ),
                    date_to=date_to,
                )

            self.cashdesk_report_org1 = self.cashdesk_report(
                database=settings.mssql_database1,
                date_from=date_from,
                date_to=date_to,
            )
            self.cashdesk_report_org1_lastyear = self.cashdesk_report(
                database=settings.mssql_database1,
                date_from=date_from - relativedelta(years=1),
                date_to=date_to - relativedelta(years=1),
            )
            self.client_count_totals_org1 = self.client_count_totals_period(
                database=settings.mssql_database1,
                org=self.org1[0],
                org_name=self.org1[1],
                date_from=date_from,
                date_to=date_to,
            )
        if self.org2:
            self.bars_srv.set_database(settings.mssql_database2)
            with self.bars_srv as connect:
                self.itog_report_org2 = functions.get_total_report(
                    connect=connect,
                    org=self.org2[0],
                    org_name=self.org2[1],
                    date_from=date_from,
                    date_to=date_to,
                    is_legacy_database=True,
                )

            self.cashdesk_report_org2 = self.cashdesk_report(
                database=settings.mssql_database2,
                date_from=date_from,
                date_to=date_to,
            )
            self.client_count_totals_org2 = self.client_count_totals_period(
                database=settings.mssql_database2,
                org=self.org2[0],
                org_name=self.org2[1],
                date_from=date_from,
                date_to=date_to,
            )

    async def run_report(self, date_from, date_to, use_yadisk: bool = False):
        self.path_list = []
        self.sms_report_list = []

        period = []
        while True:
            period.append(date_from)
            if date_from + timedelta(1) == date_to:
                break
            else:
                date_from = date_from + timedelta(1)

        # Поиск новых услуг
        for report_name in ("GoogleReport", "PlatAgentReport"):
            self._settings_service.choose_db("Aquapark_Ulyanovsk")
            new_tariffs = await self._settings_service.get_new_tariff(report_name)
            if new_tariffs:
                error_message = f"Найдены нераспределенные тарифы в отчете {report_name}: {new_tariffs}"
                logger.error(error_message)
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=error_message,
                )

        for date in period:
            date_from = date
            date_to = date + timedelta(1)
            await self.load_report(date_from, date_to)

            self.orgs_dict = (
                await self._report_config_service.get_report_elements_with_groups(
                    "GoogleReport"
                )
            )
            self.itogreport_group_dict = (
                await self._report_config_service.get_report_elements_with_groups(
                    "ItogReport"
                )
            )
            self.agent_dict = (
                await self._report_config_service.get_report_elements_with_groups(
                    "PlatAgentReport"
                )
            )
            await self.save_reports(date_from)

        # Отправка в яндекс диск
        if use_yadisk:
            self.path_list = filter(lambda x: x is not None, self.path_list)
            self._yandex_repo.sync_to_yadisk(
                self.path_list, settings.yadisk_token, date_from
            )
            self.path_list = []


def get_legacy_service() -> BarsicReport2Service:
    bars_srv = MsSqlDatabase(
        server=settings.mssql_server,
        user=settings.mssql_user,
        password=settings.mssql_pwd,
    )
    rk_srv = MsSqlDatabase(
        server=settings.mssql_server_rk,
        user=settings.mssql_user_rk,
        password=settings.mssql_pwd_rk,
    )
    return BarsicReport2Service(
        bars_srv=bars_srv,
        rk_srv=rk_srv,
    )
