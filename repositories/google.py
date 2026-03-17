import logging
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

import gspread
import gspread_formatting as gf
from openpyxl.utils import get_column_letter

from core.settings import settings

logger = logging.getLogger(__name__)


class GoogleRepository:
    def __init__(self):
        self._service_account_config = settings.google_api_settings.google_service_account_config

    def _get_client(self):
        return gspread.service_account_from_dict(self._service_account_config)

    @staticmethod
    def _share_document(google_doc) -> None:
        google_all_read = str(settings.google_api_settings.google_all_read).strip().lower()
        if google_all_read in {"1", "true", "yes"}:
            google_doc.share(None, "anyone", "reader")

        for email in settings.google_api_settings.google_writer_list.split(","):
            if not email:
                continue
            google_doc.share(email, "user", "writer")

        for email in settings.google_api_settings.google_reader_list.split(","):
            if not email:
                continue
            google_doc.share(email, "user", "reader")

    @staticmethod
    def _to_number(value: int | float | Decimal) -> int | float:
        if isinstance(value, Decimal):
            if value == value.to_integral():
                return int(value)
            return float(value)
        return value

    def save_attendance_report(
        self,
        report: dict,
        date_from: datetime,
        date_to: datetime,
        google_doc_id: str | None = None,
    ) -> tuple[str, str]:
        level1_order = [
            "тарифы (посещение)",
            "плавание",
            "гашение в барсе (посещение)",
            "товары (покупки)",
            "онлайн продукты (insales) (покупки)",
        ]
        level2_order = {
            "тарифы (посещение)": ["частные гости", "корпоративные гости", "промо"],
            "плавание": ["плавание"],
            "гашение в барсе (посещение)": ["частные гости", "корпоративные гости", "гости отеля", "промо"],
            "товары (покупки)": ["частные гости", "корпоративные гости"],
            "онлайн продукты (insales) (покупки)": ["частные гости", "корпоративные гости", "промо"],
        }
        level2_aliases = {
            "корп": "корпоративные гости",
            "корп гости": "корпоративные гости",
            "корпоративные гости": "корпоративные гости",
        }
        swimming_section = "плавание"
        attendance_sections = {"тарифы (посещение)", "гашение в барсе (посещение)"}
        manual_fill_sections = {"онлайн продукты (insales) (покупки)"}

        def normalize(text: str | None) -> str:
            if text is None:
                return ""
            return " ".join(str(text).replace("\xa0", " ").split()).strip().lower()

        def title_last(text: str | None) -> str:
            parts = [part.strip() for part in str(text or "").split("/") if part.strip()]
            if not parts:
                return ""
            return " ".join(parts[-1].replace("\xa0", " ").split()).strip()

        def canonical_level2(text: str) -> str:
            normalized = normalize(text)
            return level2_aliases.get(normalized, normalized)

        def to_date(value: date | datetime | str) -> date | None:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                try:
                    return datetime.fromisoformat(value).date()
                except ValueError:
                    return None
            return None

        def get_report_data(report_item) -> dict:
            if hasattr(report_item, "report_data"):
                return report_item.report_data or {}
            if isinstance(report_item, dict):
                return report_item.get("report_data", report_item)
            return {}

        def order_level1(level1_key: str) -> tuple[int, str]:
            try:
                return (level1_order.index(level1_key), level1_key)
            except ValueError:
                return (len(level1_order), level1_key)

        def order_level2(level1_key: str, level2_key: str) -> tuple[int, str]:
            expected_order = level2_order.get(level1_key, [])
            try:
                return (expected_order.index(level2_key), level2_key)
            except ValueError:
                return (len(expected_order), level2_key)

        metrics_tree: dict[str, dict[str, set[str]]] = {}
        level1_titles: dict[str, str] = {}
        level2_titles: dict[tuple[str, str], str] = {}
        level3_titles: dict[tuple[str, str, str], str] = {}
        reports_by_date: dict[date, dict] = {}
        swim_metric_keys: set[tuple[str, str, str]] = set()
        swim_title = "Плавание"

        for day_key, day_value in (report or {}).items():
            day = to_date(day_key)
            if day is None:
                continue

            day_data = get_report_data(day_value)
            reports_by_date[day] = day_data
            if not isinstance(day_data, dict):
                continue

            for h1, h2_data in day_data.items():
                if not isinstance(h2_data, dict):
                    continue

                h1_title = title_last(h1)
                h1_key = normalize(h1_title)
                if "количество посещений" in h1_key:
                    continue

                if h1_key == swimming_section:
                    for h2, h3_data in h2_data.items():
                        if not isinstance(h3_data, dict):
                            continue
                        h2_key = canonical_level2(title_last(h2))
                        for h3, value in h3_data.items():
                            if not isinstance(value, int | float | Decimal):
                                continue
                            h3_title = title_last(h3)
                            h3_key = normalize(h3_title)
                            swim_metric_keys.add((h1_key, h2_key, h3_key))
                            if h3_key == swimming_section:
                                swim_title = h3_title
                    continue

                level1_titles.setdefault(h1_key, h1_title)
                level1_metrics = metrics_tree.setdefault(h1_key, {})
                for h2, h3_data in h2_data.items():
                    if not isinstance(h3_data, dict):
                        continue

                    h2_title = title_last(h2)
                    h2_key = canonical_level2(h2_title)
                    level2_titles.setdefault((h1_key, h2_key), h2_title)
                    level2_metrics = level1_metrics.setdefault(h2_key, set())
                    for h3, value in h3_data.items():
                        if not isinstance(value, int | float | Decimal):
                            continue
                        h3_title = title_last(h3)
                        h3_key = normalize(h3_title)
                        level3_titles.setdefault((h1_key, h2_key, h3_key), h3_title)
                        level2_metrics.add(h3_key)

        ordered_metrics: list[tuple[str, str, str]] = []
        for level1_key in sorted(metrics_tree.keys(), key=order_level1):
            level2_map = metrics_tree[level1_key]
            for level2_key in sorted(level2_map.keys(), key=lambda item: order_level2(level1_key, item)):
                level3_keys = sorted(
                    level2_map[level2_key],
                    key=lambda item: normalize(level3_titles.get((level1_key, level2_key, item), item)),
                )
                for level3_key in level3_keys:
                    ordered_metrics.append((level1_key, level2_key, level3_key))

        attendance_metrics = [metric for metric in ordered_metrics if metric[0] in attendance_sections]
        purchase_metrics = [metric for metric in ordered_metrics if metric[0] not in attendance_sections]

        attendance_start_col = 3
        swimming_col = None
        swim_order_index = level1_order.index(swimming_section)
        swim_insert_pos = sum(1 for metric in attendance_metrics if order_level1(metric[0])[0] < swim_order_index)

        metric_column_map: dict[tuple[str, str, str], int] = {}
        current_col = attendance_start_col
        for i, metric in enumerate(attendance_metrics):
            if swim_metric_keys and i == swim_insert_pos:
                swimming_col = current_col
                current_col += 1
            metric_column_map[metric] = current_col
            current_col += 1
        if swim_metric_keys and swim_insert_pos >= len(attendance_metrics):
            swimming_col = current_col
            current_col += 1

        attendance_end_col = current_col - 1 if current_col > attendance_start_col else attendance_start_col - 1
        sales_col = attendance_end_col + 1 if attendance_end_col >= attendance_start_col else attendance_start_col
        purchase_start_col = sales_col + 1
        purchase_end_col = purchase_start_col + len(purchase_metrics) - 1
        last_col = max(2, sales_col, purchase_end_col, swimming_col or 0)

        for i, metric in enumerate(purchase_metrics):
            metric_column_map[metric] = purchase_start_col + i

        level1_by_col = {col: metric[0] for metric, col in metric_column_map.items()}
        period_end = (date_to - timedelta(days=1)).date() if date_to.time() == datetime.min.time() else date_to.date()
        if period_end < date_from.date():
            period_end = date_from.date()
        days_count = (period_end - date_from.date()).days + 1
        total_rows = 4 + days_count
        last_col_letter = get_column_letter(last_col)
        manual_metric_columns = sorted(col for metric, col in metric_column_map.items() if metric[0] in manual_fill_sections)

        report_name = "Отчет по количеству в разрезе дня"
        period_name = date_from.strftime("%Y-%m")
        client = self._get_client()
        google_doc = None
        if google_doc_id is not None:
            try:
                google_doc = client.open_by_key(google_doc_id)
            except Exception:
                logger.exception("Не удалось открыть attendance Google-документ по id=%s. Создаем новый.", google_doc_id)

        if google_doc is None:
            google_doc = client.create(f"Отчет по количеству в разрезе дня ({period_name})")
            self._share_document(google_doc)

        worksheet = google_doc.get_worksheet(0)
        existing_manual_values_by_day: dict[str, dict[int, str]] = {}
        if google_doc_id is not None and manual_metric_columns:
            existing_values = worksheet.get_all_values()
            for row_values in existing_values[4:]:
                if not row_values:
                    continue
                day_value = str(row_values[0]).strip()
                if not day_value:
                    continue
                day_manual_values = {}
                for col in manual_metric_columns:
                    if len(row_values) < col:
                        continue
                    value = row_values[col - 1]
                    if value != "":
                        day_manual_values[col] = value
                if day_manual_values:
                    existing_manual_values_by_day[day_value] = day_manual_values

        worksheet.update_title(str(date_from.year))
        worksheet.clear()
        worksheet.resize(rows=total_rows, cols=last_col)

        matrix: list[list[Any]] = [["" for _ in range(last_col)] for _ in range(total_rows)]
        matrix[0][0] = report_name
        matrix[1][0] = "Дни/Продукт"
        matrix[1][1] = "Количество посещений"
        matrix[1][sales_col - 1] = "Количество продаж"
        if swimming_col is not None:
            matrix[1][swimming_col - 1] = swim_title

        grouped_metrics = sorted(metric_column_map.items(), key=lambda item: item[1])
        if grouped_metrics:
            first_metric, first_col = grouped_metrics[0]
            start_col = first_col
            prev_col = first_col
            current_level1 = first_metric[0]
            for metric, col in grouped_metrics[1:]:
                if metric[0] != current_level1 or col != prev_col + 1:
                    matrix[1][start_col - 1] = level1_titles.get(current_level1, current_level1)
                    start_col = col
                    current_level1 = metric[0]
                prev_col = col
            matrix[1][start_col - 1] = level1_titles.get(current_level1, current_level1)

            first_metric, first_col = grouped_metrics[0]
            start_col = first_col
            prev_col = first_col
            current_pair = (first_metric[0], first_metric[1])
            for metric, col in grouped_metrics[1:]:
                pair = (metric[0], metric[1])
                if pair != current_pair or col != prev_col + 1:
                    matrix[2][start_col - 1] = level2_titles.get(current_pair, current_pair[1])
                    start_col = col
                    current_pair = pair
                prev_col = col
            matrix[2][start_col - 1] = level2_titles.get(current_pair, current_pair[1])

        for metric in attendance_metrics:
            col = metric_column_map[metric]
            matrix[3][col - 1] = level3_titles.get(metric, metric[2])
        for metric in purchase_metrics:
            col = metric_column_map[metric]
            matrix[3][col - 1] = level3_titles.get(metric, metric[2])

        current_day = date_from.date()
        row = 5
        while current_day <= period_end:
            day_as_str = current_day.strftime("%d.%m.%Y")
            matrix[row - 1][0] = day_as_str

            day_data = reports_by_date.get(current_day, {})
            customer_count = None
            swim_value: int | float = 0

            if isinstance(day_data, dict):
                for h1, h2_data in day_data.items():
                    if not isinstance(h2_data, dict):
                        continue
                    h1_key = normalize(title_last(h1))
                    for h2, h3_data in h2_data.items():
                        if not isinstance(h3_data, dict):
                            continue
                        h2_key = canonical_level2(title_last(h2))
                        for h3, value in h3_data.items():
                            if not isinstance(value, int | float | Decimal):
                                continue
                            if "количество посещений" in h1_key:
                                customer_count = value
                                continue
                            h3_key = normalize(title_last(h3))
                            metric_key = (h1_key, h2_key, h3_key)
                            if metric_key in swim_metric_keys:
                                swim_value += self._to_number(value)
                                continue
                            if metric_key[0] in manual_fill_sections:
                                continue
                            col = metric_column_map.get(metric_key)
                            if col is not None:
                                matrix[row - 1][col - 1] = self._to_number(value)

            if swimming_col is not None:
                matrix[row - 1][swimming_col - 1] = self._to_number(swim_value)

            if attendance_end_col >= attendance_start_col:
                matrix[row - 1][1] = (
                    f"=SUM({get_column_letter(attendance_start_col)}{row}:{get_column_letter(attendance_end_col)}{row})"
                )
            elif customer_count is not None:
                matrix[row - 1][1] = self._to_number(customer_count)
            else:
                matrix[row - 1][1] = 0

            if purchase_metrics:
                matrix[row - 1][sales_col - 1] = (
                    f"=SUM({get_column_letter(purchase_start_col)}{row}:{get_column_letter(purchase_end_col)}{row})"
                )
            else:
                matrix[row - 1][sales_col - 1] = 0

            existed_manual_values = existing_manual_values_by_day.get(day_as_str)
            if existed_manual_values:
                for col, value in existed_manual_values.items():
                    matrix[row - 1][col - 1] = value

            row += 1
            current_day += timedelta(days=1)

        worksheet.update(matrix, f"A1:{last_col_letter}{total_rows}", raw=False)

        merge_requests = []
        vertical_ranges: list[tuple[int, int, int, int]] = [(4, 4, 2, last_col)]

        def add_merge(start_row: int, end_row: int, start_col: int, end_col: int) -> None:
            if start_row == end_row and start_col == end_col:
                return
            merge_requests.append(
                {
                    "mergeCells": {
                        "range": {
                            "sheetId": worksheet.id,
                            "startRowIndex": start_row - 1,
                            "endRowIndex": end_row,
                            "startColumnIndex": start_col - 1,
                            "endColumnIndex": end_col,
                        },
                        "mergeType": "MERGE_ALL",
                    }
                }
            )

        add_merge(1, 1, 1, last_col)

        if grouped_metrics:
            first_metric, first_col = grouped_metrics[0]
            start_col = first_col
            prev_col = first_col
            current_level1 = first_metric[0]
            for metric, col in grouped_metrics[1:]:
                if metric[0] != current_level1 or col != prev_col + 1:
                    add_merge(2, 2, start_col, prev_col)
                    start_col = col
                    current_level1 = metric[0]
                prev_col = col
            add_merge(2, 2, start_col, prev_col)

            first_metric, first_col = grouped_metrics[0]
            start_col = first_col
            prev_col = first_col
            current_pair = (first_metric[0], first_metric[1])
            for metric, col in grouped_metrics[1:]:
                pair = (metric[0], metric[1])
                if pair != current_pair or col != prev_col + 1:
                    add_merge(3, 3, start_col, prev_col)
                    start_col = col
                    current_pair = pair
                prev_col = col
            add_merge(3, 3, start_col, prev_col)

        add_merge(2, 4, 1, 1)
        add_merge(2, 4, 2, 2)
        vertical_ranges.append((2, 4, 2, 2))
        add_merge(2, 4, sales_col, sales_col)
        vertical_ranges.append((2, 4, sales_col, sales_col))
        if swimming_col is not None:
            add_merge(2, 4, swimming_col, swimming_col)
            vertical_ranges.append((2, 4, swimming_col, swimming_col))

        level3_width = 40
        merged_header_width = int(level3_width * 1.2)
        merged_header_columns = {2, sales_col}
        if swimming_col is not None:
            merged_header_columns.add(swimming_col)
        hotel_level2_width = int(level3_width * 1.5)
        corp_level2_width = int(level3_width * 1.2)
        hotel_level2_columns = sorted(col for metric, col in metric_column_map.items() if metric[1] == "гости отеля")
        corp_level2_columns = sorted(col for metric, col in metric_column_map.items() if metric[1] == "корпоративные гости")

        resize_requests = [
            {
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet.id,
                        "startRowIndex": 0,
                        "endRowIndex": total_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": last_col,
                    },
                    "cell": {
                        "userEnteredFormat": {},
                    },
                    "fields": "userEnteredFormat",
                }
            },
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": worksheet.id,
                        "gridProperties": {
                            "frozenRowCount": 4,
                        },
                    },
                    "fields": "gridProperties.frozenRowCount",
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "COLUMNS",
                        "startIndex": 0,
                        "endIndex": 1,
                    },
                    "properties": {"pixelSize": 120},
                    "fields": "pixelSize",
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "COLUMNS",
                        "startIndex": 1,
                        "endIndex": last_col,
                    },
                    "properties": {"pixelSize": level3_width},
                    "fields": "pixelSize",
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "ROWS",
                        "startIndex": 0,
                        "endIndex": 1,
                    },
                    "properties": {"pixelSize": 32},
                    "fields": "pixelSize",
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "ROWS",
                        "startIndex": 1,
                        "endIndex": 2,
                    },
                    "properties": {"pixelSize": 34},
                    "fields": "pixelSize",
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "ROWS",
                        "startIndex": 2,
                        "endIndex": 3,
                    },
                    "properties": {"pixelSize": 41},
                    "fields": "pixelSize",
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "ROWS",
                        "startIndex": 3,
                        "endIndex": 4,
                    },
                    "properties": {"pixelSize": 168},
                    "fields": "pixelSize",
                }
            },
        ]

        resize_requests.extend(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "COLUMNS",
                        "startIndex": col - 1,
                        "endIndex": col,
                    },
                    "properties": {"pixelSize": merged_header_width},
                    "fields": "pixelSize",
                }
            }
            for col in sorted(merged_header_columns)
        )
        resize_requests.extend(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "COLUMNS",
                        "startIndex": col - 1,
                        "endIndex": col,
                    },
                    "properties": {"pixelSize": hotel_level2_width},
                    "fields": "pixelSize",
                }
            }
            for col in hotel_level2_columns
        )
        resize_requests.extend(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet.id,
                        "dimension": "COLUMNS",
                        "startIndex": col - 1,
                        "endIndex": col,
                    },
                    "properties": {"pixelSize": corp_level2_width},
                    "fields": "pixelSize",
                }
            }
            for col in corp_level2_columns
        )
        resize_requests.extend(
            {
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet.id,
                        "startRowIndex": start_row - 1,
                        "endRowIndex": end_row,
                        "startColumnIndex": start_col - 1,
                        "endColumnIndex": end_col,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "textRotation": {
                                "angle": 90,
                            }
                        }
                    },
                    "fields": "userEnteredFormat.textRotation",
                }
            }
            for start_row, end_row, start_col, end_col in vertical_ranges
        )

        google_doc.batch_update({"requests": [*merge_requests, *resize_requests]})

        border = gf.Border(style="SOLID")
        h1_fmt = gf.CellFormat(
            textFormat=gf.TextFormat(fontSize=18, bold=True),
            horizontalAlignment="CENTER",
            verticalAlignment="MIDDLE",
        )
        header_h2_h3_fmt = gf.CellFormat(
            textFormat=gf.TextFormat(fontSize=11, bold=True),
            horizontalAlignment="CENTER",
            verticalAlignment="MIDDLE",
            wrapStrategy="WRAP",
            borders=gf.Borders(top=border, right=border, bottom=border, left=border),
        )
        header_h4_fmt = gf.CellFormat(
            textFormat=gf.TextFormat(fontSize=9, bold=True),
            horizontalAlignment="CENTER",
            verticalAlignment="BOTTOM",
            wrapStrategy="WRAP",
            borders=gf.Borders(top=border, right=border, bottom=border, left=border),
        )
        data_fmt = gf.CellFormat(
            textFormat=gf.TextFormat(fontSize=9, bold=False),
            horizontalAlignment="CENTER",
            verticalAlignment="MIDDLE",
            borders=gf.Borders(top=border, right=border, bottom=border, left=border),
        )
        day_header_fill = gf.CellFormat(backgroundColor=gf.Color.fromHex("#f7cb4d"))
        total_header_fill = gf.CellFormat(backgroundColor=gf.Color.fromHex("#ffff00"))
        default_group_fill = gf.CellFormat(backgroundColor=gf.Color.fromHex("#e2e2e2"))
        group_fills = {
            "тарифы (посещение)": gf.CellFormat(backgroundColor=gf.Color.fromHex("#f7cb4d")),
            "плавание": gf.CellFormat(backgroundColor=gf.Color.fromHex("#9bd4f5")),
            "гашение в барсе (посещение)": gf.CellFormat(backgroundColor=gf.Color.fromHex("#f4a6a6")),
            "товары (покупки)": gf.CellFormat(backgroundColor=gf.Color.fromHex("#b7d7a8")),
            "онлайн продукты (insales) (покупки)": gf.CellFormat(backgroundColor=gf.Color.fromHex("#f9cb9c")),
        }

        format_ranges = [
            ("A1:A1", h1_fmt),
            (f"A2:{last_col_letter}3", header_h2_h3_fmt),
            (f"A4:{last_col_letter}4", header_h4_fmt),
        ]
        if total_rows >= 5:
            format_ranges.append((f"A5:{last_col_letter}{total_rows}", data_fmt))

        for col in range(1, last_col + 1):
            col_letter = get_column_letter(col)
            if col == 1:
                fill_fmt = day_header_fill
            elif col in {2, sales_col}:
                fill_fmt = total_header_fill
            elif swimming_col is not None and col == swimming_col:
                fill_fmt = group_fills.get(swimming_section, default_group_fill)
            else:
                fill_fmt = group_fills.get(level1_by_col.get(col), default_group_fill)
            format_ranges.append((f"{col_letter}2:{col_letter}4", fill_fmt))

        gf.format_cell_ranges(worksheet, format_ranges)
        return google_doc.url, google_doc.id


def get_google_repo() -> GoogleRepository:
    return GoogleRepository()
