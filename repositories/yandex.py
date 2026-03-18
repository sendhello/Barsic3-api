import logging
import os
from datetime import date, datetime, timedelta
from decimal import Decimal

from fastapi.exceptions import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from starlette import status
from yadisk import AsyncYaDisk, YaDisk
from yadisk.objects import SyncResourceLinkObject

from core.settings import settings

logger = logging.getLogger(__name__)


class ReportStyle:
    h1 = Font(
        name="Times New Roman",
        size=18,
        bold=True,
        italic=False,
        vertAlign=None,
        underline="none",
        strike=False,
        color="FF000000",
    )
    h2 = Font(
        name="Times New Roman",
        size=14,
        bold=True,
        italic=False,
        vertAlign=None,
        underline="none",
        strike=False,
        color="FF000000",
    )
    h3 = Font(
        name="Times New Roman",
        size=11,
        bold=True,
        italic=False,
        vertAlign=None,
        underline="none",
        strike=False,
        color="FF000000",
    )
    font = Font(
        name="Times New Roman",
        size=9,
        bold=False,
        italic=False,
        vertAlign=None,
        underline="none",
        strike=False,
        color="FF000000",
    )
    font_bold = Font(
        name="Times New Roman",
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline="none",
        strike=False,
        color="FF000000",
    )
    font_red = Font(
        name="Times New Roman",
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline="none",
        strike=False,
        color="FFFF0000",
    )
    align_bottom = Alignment(
        horizontal="general",
        vertical="bottom",
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0,
    )
    border = Border(
        left=Side(border_style="thin", color="FF000000"),
        right=Side(border_style="thin", color="FF000000"),
        top=Side(border_style="thin", color="FF000000"),
        bottom=Side(border_style="thin", color="FF000000"),
        diagonal=Side(border_style="thin", color="FF000000"),
        diagonal_direction=0,
        outline=Side(border_style="thin", color="FF000000"),
        vertical=Side(border_style="thin", color="FF000000"),
        horizontal=Side(border_style="thin", color="FF000000"),
    )
    border_top_bottom = Border(
        bottom=Side(border_style="thin", color="FF000000"),
        top=Side(border_style="thin", color="FF000000"),
    )
    border_right = Border(right=Side(border_style="thin", color="FF000000"))
    border_left = Border(left=Side(border_style="thin", color="FF000000"))
    border_top = Border(top=Side(border_style="thin", color="FF000000"))
    border_left_top = Border(
        top=Side(border_style="thin", color="FF000000"),
        left=Side(border_style="thin", color="FF000000"),
    )
    border_right_top = Border(
        top=Side(border_style="thin", color="FF000000"),
        right=Side(border_style="thin", color="FF000000"),
    )
    fill = PatternFill(fill_type="solid", start_color="c1c1c1", end_color="c2c2c2")
    align_top = Alignment(
        horizontal="general",
        vertical="top",
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0,
    )
    align_left = Alignment(
        horizontal="left",
        vertical="bottom",
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0,
    )


class YandexRepository:
    def __init__(self, token: str):
        self._yadisk = YaDisk(token=token)
        self._async_yadisk = AsyncYaDisk(token=token)

    def save_purchased_goods_report(
        self,
        report,
        date_from,
        date_to,
        goods: list[str] = None,
        hide_zero=False,
    ):
        """Сохраняет отчет по количеству клиентов за день в Excel"""

        column = ["", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Суммы трат в аквазоне"
        # шрифты
        ws["C1"].font = ReportStyle.h1
        # выравнивание
        ws["C1"].alignment = ReportStyle.align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 8
        ws.column_dimensions["B"].width = 1 / 7 * 8
        ws.column_dimensions["C"].width = 1 / 7 * 80
        ws.column_dimensions["D"].width = 1 / 7 * 8
        ws.column_dimensions["E"].width = 1 / 7 * 88
        ws.column_dimensions["F"].width = 1 / 7 * 8
        ws.column_dimensions["G"].width = 1 / 7 * 24
        ws.column_dimensions["H"].width = 1 / 7 * 8
        ws.column_dimensions["I"].width = 1 / 7 * 80
        ws.column_dimensions["J"].width = 1 / 7 * 8
        ws.column_dimensions["K"].width = 1 / 7 * 144
        ws.column_dimensions["L"].width = 1 / 7 * 144
        ws.column_dimensions["M"].width = 1 / 7 * 8

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[3] + next_row()] = "Отчет по купленным товарам"
        ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=12)
        ws[column[1] + next_row()] = ""
        ws[column[3] + next_row()] = f"По товарам: {', '.join(goods)}"
        ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=12)
        ws[column[3] + self.row].font = ReportStyle.font
        ws[column[3] + self.row].alignment = ReportStyle.align_top
        ws[column[1] + next_row()] = ""

        if date_from == date_to - timedelta(days=1):
            ws[column[3] + next_row()] = "За:"
            ws[column[3] + self.row].font = ReportStyle.font
            ws[column[3] + self.row].alignment = ReportStyle.align_top
            ws[column[5] + self.row] = date_from.strftime("%d.%m.%Y")
            ws[column[5] + self.row].font = ReportStyle.font_bold
            ws[column[5] + self.row].alignment = ReportStyle.align_top
        else:
            ws[column[3] + next_row()] = "За период с:"
            ws[column[3] + self.row].font = ReportStyle.font
            ws[column[3] + self.row].alignment = ReportStyle.align_top
            ws[column[5] + self.row] = date_from.strftime("%d.%m.%Y")
            ws[column[5] + self.row].font = ReportStyle.font_bold
            ws[column[5] + self.row].alignment = ReportStyle.align_top
            ws[column[7] + self.row] = "По:"
            ws[column[7] + self.row].font = ReportStyle.font
            ws[column[7] + self.row].alignment = ReportStyle.align_top
            ws[column[9] + self.row] = (date_to - timedelta(days=1)).strftime("%d.%m.%Y")
            ws[column[9] + self.row].font = ReportStyle.font_bold
            ws[column[9] + self.row].alignment = ReportStyle.align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.font
            ws[column[10] + self.row].font = ReportStyle.font
            ws[column[12] + self.row].font = ReportStyle.font
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[12] + self.row].alignment = ReportStyle.align_top
            b = 2
            while b <= 13:
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        def merge_table_h3():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.h3
            ws[column[10] + self.row].font = ReportStyle.h3
            ws[column[12] + self.row].font = ReportStyle.h3
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[12] + self.row].alignment = ReportStyle.align_top
            ws[column[2] + self.row].border = ReportStyle.border_left
            ws[column[13] + self.row].border = ReportStyle.border_right

        def merge_table_h2():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.h2
            ws[column[10] + self.row].font = ReportStyle.h2
            ws[column[12] + self.row].font = ReportStyle.h2
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[12] + self.row].alignment = ReportStyle.align_top
            ws[column[2] + self.row].border = ReportStyle.border_left
            ws[column[13] + self.row].border = ReportStyle.border_right

        def merge_width_h2():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.h2
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            b = 2
            while b <= 13:
                if b == 2:
                    ws[column[b] + self.row].border = ReportStyle.border_left_top
                elif b == 13:
                    ws[column[b] + self.row].border = ReportStyle.border_right_top
                else:
                    ws[column[b] + self.row].border = ReportStyle.border_top
                b += 1

        def merge_width_h3():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.h3
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            b = 2
            while b <= 13:
                if b == 2:
                    ws[column[b] + self.row].border = ReportStyle.border_left
                elif b == 13:
                    ws[column[b] + self.row].border = ReportStyle.border_right
                b += 1

        ws[column[2] + next_row()] = "Наименование"
        ws[column[10] + self.row] = "Количество"
        ws[column[12] + self.row] = "Сумма"
        merge_table()
        ws[column[2] + self.row].font = ReportStyle.h3
        ws[column[10] + self.row].font = ReportStyle.h3
        ws[column[12] + self.row].font = ReportStyle.h3
        ws[column[2] + self.row].alignment = ReportStyle.align_top
        ws[column[10] + self.row].alignment = ReportStyle.align_top
        ws[column[12] + self.row].alignment = ReportStyle.align_top

        for line in report:
            if hide_zero and line.summ == Decimal(0):
                continue

            ws[column[2] + next_row()] = line.name[8:] if line.name.startswith("Долг за") else line.name
            ws[column[10] + self.row] = line.count
            ws[column[12] + self.row] = line.summ
            ws[column[12] + self.row].number_format = "#,##0.00 ₽"
            merge_table()

        ws[column[2] + next_row()] = "Итого"
        ws[column[10] + self.row] = sum(line.count for line in report if line.name not in goods)
        ws[column[12] + self.row] = sum(line.summ for line in report if line.name not in goods)
        ws[column[12] + self.row].number_format = "#,##0.00 ₽"
        merge_table_h2()
        ws[column[2] + self.row].alignment = ReportStyle.align_bottom
        ws[column[10] + self.row].alignment = ReportStyle.align_bottom
        ws[column[12] + self.row].alignment = ReportStyle.align_bottom
        b = 2
        while b <= 13:
            ws[column[b] + self.row].border = ReportStyle.border_top_bottom
            b += 1
        end_line = int(self.row)

        # раскрвшивание фона для заголовков
        i = 2
        while i <= 13:
            ws[column[i] + "6"].fill = ReportStyle.fill
            i += 1

        # обводка
        # ws['A3'].border = ReportStyle.border

        # вручную устанавливаем высоту первой строки
        rd = ws.row_dimensions[1]
        rd.height = 21.75

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1

        # Высота строк
        ws.row_dimensions[2].height = 5.25
        ws.row_dimensions[4].height = 6.75
        ws.row_dimensions[end_line].height = 30.75

        # выравнивание столбца
        for cellObj in ws["A2:A5"]:
            for cell in cellObj:
                ws[cell.coordinate].alignment = ReportStyle.align_left

        # перетягивание ячеек
        # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
        # dims = {}
        # for row in ws.rows:
        #     for cell in row:
        #         if cell.value:
        #             dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        # for col, value in dims.items():
        #     # value * коэфициент
        #     ws.column_dimensions[col].width = value * 1.5

        if date_from == date_to - timedelta(days=1):
            date_ = datetime.strftime(date_from, "%Y-%m-%d")
        else:
            date_ = (
                f"{datetime.strftime(date_from, '%Y-%m-%d')} - "
                f"{datetime.strftime(date_to - timedelta(days=1), '%Y-%m-%d')}"
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f" Отчет по купленным товарам ({datetime.now().timestamp()}).xlsx"
        )
        logger.info(f"Сохранение отчета по купленным товарам за {date_} в {path}")
        path = self.create_path(path, date_from)
        self.save_file(path, wb)
        return path

    def save_attendance_report(
        self,
        report: dict,
        date_from: datetime,
        date_to: datetime,
    ) -> str:
        """Сохраняет отчет по посещаемости в Excel (генерация с нуля)."""

        level1_order = [
            "тарифы (посещение)",
            "плавание",
            "гашение в барсе (посещение)",
            "товары (покупки)",
            "онлайн продукты (insales) (покупки)",
        ]
        level2_order = {
            "тарифы (посещение)": ["частные гости", "корпоративные гости", "промо"],
            "плавание": ["плавание"],
            "гашение в барсе (посещение)": ["частные гости", "корпоративные гости", "гости отеля", "промо"],
            "товары (покупки)": ["частные гости", "корпоративные гости"],
            "онлайн продукты (insales) (покупки)": ["частные гости", "корпоративные гости", "промо"],
        }
        level2_aliases = {
            "корп": "корпоративные гости",
            "корп гости": "корпоративные гости",
            "корпоративные гости": "корпоративные гости",
        }
        swimming_section = "плавание"
        attendance_sections = {"тарифы (посещение)", "гашение в барсе (посещение)"}

        def normalize(text: str | None) -> str:
            if text is None:
                return ""
            return " ".join(str(text).replace("\xa0", " ").split()).strip().lower()

        def title_last(text: str | None) -> str:
            parts = [part.strip() for part in str(text or "").split("/") if part.strip()]
            if not parts:
                return ""
            return " ".join(parts[-1].replace("\xa0", " ").split()).strip()

        def canonical_level2(text: str) -> str:
            normalized = normalize(text)
            return level2_aliases.get(normalized, normalized)

        def to_date(value: date | datetime | str) -> date | None:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                try:
                    return datetime.fromisoformat(value).date()
                except ValueError:
                    return None
            return None

        def get_report_data(report_item) -> dict:
            if hasattr(report_item, "report_data"):
                return report_item.report_data or {}
            if isinstance(report_item, dict):
                return report_item.get("report_data", report_item)
            return {}

        def order_level1(level1_key: str) -> tuple[int, str]:
            try:
                return (level1_order.index(level1_key), level1_key)
            except ValueError:
                return (len(level1_order), level1_key)

        def order_level2(level1_key: str, level2_key: str) -> tuple[int, str]:
            expected_order = level2_order.get(level1_key, [])
            try:
                return (expected_order.index(level2_key), level2_key)
            except ValueError:
                return (len(expected_order), level2_key)

        metrics_tree: dict[str, dict[str, set[str]]] = {}
        level1_titles: dict[str, str] = {}
        level2_titles: dict[tuple[str, str], str] = {}
        level3_titles: dict[tuple[str, str, str], str] = {}
        reports_by_date: dict[date, dict] = {}
        swim_metric_keys: set[tuple[str, str, str]] = set()
        swim_title = "Плавание"

        for day_key, day_value in (report or {}).items():
            day = to_date(day_key)
            if day is None:
                continue

            day_data = get_report_data(day_value)
            reports_by_date[day] = day_data

            if not isinstance(day_data, dict):
                continue
            for h1, h2_data in day_data.items():
                if not isinstance(h2_data, dict):
                    continue

                h1_title = title_last(h1)
                h1_key = normalize(h1_title)
                if "количество посещений" in h1_key:
                    continue

                if h1_key == swimming_section:
                    for h2, h3_data in h2_data.items():
                        if not isinstance(h3_data, dict):
                            continue
                        h2_key = canonical_level2(title_last(h2))
                        for h3, value in h3_data.items():
                            if not isinstance(value, int | float | Decimal):
                                continue
                            h3_title = title_last(h3)
                            h3_key = normalize(h3_title)
                            swim_metric_keys.add((h1_key, h2_key, h3_key))
                            if h3_key == swimming_section:
                                swim_title = h3_title
                    continue

                level1_titles.setdefault(h1_key, h1_title)
                level1_metrics = metrics_tree.setdefault(h1_key, {})
                for h2, h3_data in h2_data.items():
                    if not isinstance(h3_data, dict):
                        continue

                    h2_title = title_last(h2)
                    h2_key = canonical_level2(h2_title)
                    level2_titles.setdefault((h1_key, h2_key), h2_title)
                    level2_metrics = level1_metrics.setdefault(h2_key, set())
                    for h3, value in h3_data.items():
                        if not isinstance(value, int | float | Decimal):
                            continue
                        h3_title = title_last(h3)
                        h3_key = normalize(h3_title)
                        level3_titles.setdefault((h1_key, h2_key, h3_key), h3_title)
                        level2_metrics.add(h3_key)

        ordered_metrics: list[tuple[str, str, str]] = []
        for level1_key in sorted(metrics_tree.keys(), key=order_level1):
            level2_map = metrics_tree[level1_key]
            for level2_key in sorted(level2_map.keys(), key=lambda item: order_level2(level1_key, item)):
                level3_keys = sorted(
                    level2_map[level2_key],
                    key=lambda item: normalize(level3_titles.get((level1_key, level2_key, item), item)),
                )
                for level3_key in level3_keys:
                    ordered_metrics.append((level1_key, level2_key, level3_key))

        attendance_metrics = [metric for metric in ordered_metrics if metric[0] in attendance_sections]
        purchase_metrics = [metric for metric in ordered_metrics if metric[0] not in attendance_sections]

        attendance_start_col = 3
        swimming_col = None
        swim_order_index = level1_order.index(swimming_section)
        swim_insert_pos = sum(1 for metric in attendance_metrics if order_level1(metric[0])[0] < swim_order_index)

        metric_column_map: dict[tuple[str, str, str], int] = {}
        current_col = attendance_start_col
        for i, metric in enumerate(attendance_metrics):
            if swim_metric_keys and i == swim_insert_pos:
                swimming_col = current_col
                current_col += 1
            metric_column_map[metric] = current_col
            current_col += 1
        if swim_metric_keys and swim_insert_pos >= len(attendance_metrics):
            swimming_col = current_col
            current_col += 1

        attendance_end_col = current_col - 1 if current_col > attendance_start_col else attendance_start_col - 1
        sales_col = attendance_end_col + 1 if attendance_end_col >= attendance_start_col else attendance_start_col
        purchase_start_col = sales_col + 1
        purchase_end_col = purchase_start_col + len(purchase_metrics) - 1
        last_col = max(2, sales_col, purchase_end_col, swimming_col or 0)

        for i, metric in enumerate(purchase_metrics):
            metric_column_map[metric] = purchase_start_col + i
        level1_by_col = {col: metric[0] for metric, col in metric_column_map.items()}
        grouped_metric_columns = set(metric_column_map.values())

        wb = Workbook()
        ws = wb.active
        ws.title = str(date_from.year)

        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        metric_header_align = Alignment(horizontal="center", vertical="bottom", text_rotation=90, wrap_text=True)
        data_align = Alignment(horizontal="center", vertical="center")
        day_header_fill = PatternFill(fill_type="solid", start_color="FFF7CB4D", end_color="FFF7CB4D")
        total_header_fill = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
        default_group_fill = PatternFill(fill_type="solid", start_color="FFE2E2E2", end_color="FFE2E2E2")
        group_fills = {
            "тарифы (посещение)": PatternFill(fill_type="solid", start_color="FFF7CB4D", end_color="FFF7CB4D"),
            "плавание": PatternFill(fill_type="solid", start_color="FF9BD4F5", end_color="FF9BD4F5"),
            "гашение в барсе (посещение)": PatternFill(fill_type="solid", start_color="FFF4A6A6", end_color="FFF4A6A6"),
            "товары (покупки)": PatternFill(fill_type="solid", start_color="FFB7D7A8", end_color="FFB7D7A8"),
            "онлайн продукты (insales) (покупки)": PatternFill(
                fill_type="solid",
                start_color="FFF9CB9C",
                end_color="FFF9CB9C",
            ),
        }

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws.cell(row=1, column=1).value = "Отчет по количеству в разрезе дня"
        ws.cell(row=1, column=1).font = ReportStyle.h1
        ws.cell(row=1, column=1).alignment = header_align

        grouped_metrics = sorted(metric_column_map.items(), key=lambda item: item[1])
        if grouped_metrics:
            first_metric, first_col = grouped_metrics[0]
            start_col = first_col
            prev_col = first_col
            current_level1 = first_metric[0]
            for metric, col in grouped_metrics[1:]:
                if metric[0] != current_level1 or col != prev_col + 1:
                    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=prev_col)
                    ws.cell(row=2, column=start_col).value = level1_titles.get(current_level1, current_level1)
                    start_col = col
                    current_level1 = metric[0]
                prev_col = col
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=prev_col)
            ws.cell(row=2, column=start_col).value = level1_titles.get(current_level1, current_level1)

            first_metric, first_col = grouped_metrics[0]
            start_col = first_col
            prev_col = first_col
            current_pair = (first_metric[0], first_metric[1])
            for metric, col in grouped_metrics[1:]:
                pair = (metric[0], metric[1])
                if pair != current_pair or col != prev_col + 1:
                    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=prev_col)
                    ws.cell(row=3, column=start_col).value = level2_titles.get(current_pair, current_pair[1])
                    start_col = col
                    current_pair = pair
                prev_col = col
            ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=prev_col)
            ws.cell(row=3, column=start_col).value = level2_titles.get(current_pair, current_pair[1])

        for metric in attendance_metrics:
            col = metric_column_map[metric]
            ws.cell(row=4, column=col).value = level3_titles.get(metric, metric[2])
        for metric in purchase_metrics:
            col = metric_column_map[metric]
            ws.cell(row=4, column=col).value = level3_titles.get(metric, metric[2])

        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
        ws.cell(row=2, column=1).value = "Дни/Продукт"

        ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
        ws.cell(row=2, column=2).value = "Количество посещений"

        ws.merge_cells(start_row=2, start_column=sales_col, end_row=4, end_column=sales_col)
        ws.cell(row=2, column=sales_col).value = "Количество продаж"

        if swimming_col is not None:
            ws.merge_cells(start_row=2, start_column=swimming_col, end_row=4, end_column=swimming_col)
            ws.cell(row=2, column=swimming_col).value = swim_title

        special_vertical_cols = {2, sales_col}
        if swimming_col is not None:
            special_vertical_cols.add(swimming_col)

        for row in range(2, 5):
            for col in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = ReportStyle.h3 if row in {2, 3} else ReportStyle.font_bold
                if col in special_vertical_cols or (row == 4 and col in grouped_metric_columns):
                    cell.alignment = metric_header_align
                else:
                    cell.alignment = header_align
                cell.border = ReportStyle.border
                if col == 1:
                    cell.fill = day_header_fill
                elif col in {2, sales_col}:
                    cell.fill = total_header_fill
                elif swimming_col is not None and col == swimming_col:
                    cell.fill = group_fills.get(swimming_section, default_group_fill)
                else:
                    cell.fill = group_fills.get(level1_by_col.get(col), default_group_fill)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 7
        if swimming_col is not None:
            ws.column_dimensions[get_column_letter(swimming_col)].width = 7
        ws.column_dimensions[get_column_letter(sales_col)].width = 7
        for metric, col in metric_column_map.items():
            header = level3_titles.get(metric, metric[2])
            width = min(max(len(header) * 0.18 + 2, 4.5), 10)
            if metric[1] == "гости отеля":
                width = min(width * 2, 20)
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 34
        ws.row_dimensions[3].height = 34
        ws.row_dimensions[4].height = 140
        ws.freeze_panes = "A5"

        period_end = (date_to - timedelta(days=1)).date() if date_to.time() == datetime.min.time() else date_to.date()
        period_end = max(period_end, date_from.date())

        current_day = date_from.date()
        row = 5
        while current_day <= period_end:
            ws.cell(row=row, column=1).value = current_day
            ws.cell(row=row, column=1).number_format = "DD.MM.YYYY"

            day_data = reports_by_date.get(current_day, {})
            customer_count = None
            swim_value = 0

            if isinstance(day_data, dict):
                for h1, h2_data in day_data.items():
                    if not isinstance(h2_data, dict):
                        continue
                    h1_key = normalize(title_last(h1))
                    for h2, h3_data in h2_data.items():
                        if not isinstance(h3_data, dict):
                            continue
                        h2_key = canonical_level2(title_last(h2))
                        for h3, value in h3_data.items():
                            if not isinstance(value, int | float | Decimal):
                                continue
                            if "количество посещений" in h1_key:
                                customer_count = value
                                continue
                            h3_key = normalize(title_last(h3))
                            metric_key = (h1_key, h2_key, h3_key)
                            if metric_key in swim_metric_keys:
                                swim_value += value
                                continue
                            col = metric_column_map.get(metric_key)
                            if col is not None:
                                ws.cell(row=row, column=col).value = value

            if swimming_col is not None:
                ws.cell(row=row, column=swimming_col).value = swim_value

            if attendance_end_col >= attendance_start_col:
                ws.cell(
                    row=row, column=2
                ).value = (
                    f"=SUM({get_column_letter(attendance_start_col)}{row}:{get_column_letter(attendance_end_col)}{row})"
                )
            elif customer_count is not None:
                ws.cell(row=row, column=2).value = customer_count
            else:
                ws.cell(row=row, column=2).value = 0

            if purchase_metrics:
                ws.cell(
                    row=row, column=sales_col
                ).value = (
                    f"=SUM({get_column_letter(purchase_start_col)}{row}:{get_column_letter(purchase_end_col)}{row})"
                )
            else:
                ws.cell(row=row, column=sales_col).value = 0

            for col in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = ReportStyle.font
                cell.alignment = data_align
                cell.border = ReportStyle.border

            row += 1
            current_day += timedelta(days=1)

        period_name = date_from.strftime("%Y-%m")

        path = settings.local_folder + settings.report_path + period_name + " Отчет посещаемости.xlsx"
        logger.info(f"Сохранение отчета по посещаемости за {period_name} в {path}")
        path = self.create_path(path, date_from)
        self.save_file(path, wb)
        return path

    def export_payment_agent_report(self, report: dict[str, list], date_from: datetime):
        """Сохраняет отчет платежного агента в виде Excel-файла в локальную директорию"""

        table_color = PatternFill(fill_type="solid", start_color="e2e2e2", end_color="e9e9e9")

        column = ["", "A", "B", "C", "D", "E"]

        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Отчет платежного агента"
        # шрифты
        ws["A1"].font = ReportStyle.h1
        # выравнивание
        ws["A1"].alignment = ReportStyle.align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 124
        ws.column_dimensions["B"].width = 1 / 7 * 80
        ws.column_dimensions["C"].width = 1 / 7 * 24
        ws.column_dimensions["D"].width = 1 / 7 * 175
        ws.column_dimensions["E"].width = 1 / 7 * 200

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = "Отчет платежного агента по приему денежных средств"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        # шрифты
        ws[column[1] + self.row].font = ReportStyle.h1
        # выравнивание
        ws[column[1] + self.row].alignment = ReportStyle.align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f"{report['Организация'][1]}"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        ws[column[1] + self.row].font = ReportStyle.font
        ws[column[1] + self.row].alignment = ReportStyle.align_top

        ws[column[1] + next_row()] = "За период с:"
        ws[column[1] + self.row].font = ReportStyle.font
        ws[column[1] + self.row].alignment = ReportStyle.align_top
        ws[column[2] + self.row] = (report["Дата"][0]).strftime("%d.%m.%Y")
        ws[column[2] + self.row].font = ReportStyle.font_bold
        ws[column[2] + self.row].alignment = ReportStyle.align_top
        ws[column[3] + self.row] = "по"
        ws[column[3] + self.row].font = ReportStyle.font
        ws[column[3] + self.row].alignment = ReportStyle.align_top
        ws[column[4] + self.row] = (report["Дата"][1] - timedelta(1)).strftime("%d.%m.%Y")
        ws[column[4] + self.row].font = ReportStyle.font_bold
        ws[column[4] + self.row].alignment = ReportStyle.align_top

        # ТАБЛИЦА
        self.color = False

        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=4)
            ws[column[1] + self.row].font = ReportStyle.font
            ws[column[5] + self.row].font = ReportStyle.font
            ws[column[1] + self.row].alignment = ReportStyle.align_top
            ws[column[5] + self.row].alignment = ReportStyle.align_top
            ws[column[5] + self.row].number_format = "#,##0.00 ₽"
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1
            if self.color:
                b = 1
                while b < len(column):
                    ws[column[b] + self.row].fill = table_color
                    b += 1
                self.color = False
            else:
                self.color = True

        def merge_table_bold():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=4)
            ws[column[1] + self.row].font = ReportStyle.font_bold
            ws[column[5] + self.row].font = ReportStyle.font_bold
            ws[column[1] + self.row].alignment = ReportStyle.align_top
            ws[column[5] + self.row].alignment = ReportStyle.align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        ws[column[1] + next_row()] = "Наименование поставщика услуг"
        ws[column[5] + self.row] = "Сумма"
        merge_table_bold()
        # раскрашивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = ReportStyle.fill
            b += 1

        itog_sum = 0
        for line in report:
            if line not in {"Организация", "Дата", "ИТОГО"}:
                try:
                    itog_sum += report[line][1]
                    ws[column[1] + next_row()] = line
                    ws[column[5] + self.row] = report[line][1]
                    merge_table()
                except AttributeError:
                    pass

        ws[column[1] + next_row()] = "Итого"
        if itog_sum != report["ИТОГО"][1]:
            logger.error(
                f"Ошибка. Отчет платежного агента: сумма строк "
                f"({itog_sum}) не равна строке ИТОГО "
                f"({report['ИТОГО'][1]})"
            )
            logger.info(
                "Ошибка. Отчет платежного агента",
                "Ошибка. Отчет платежного агента: сумма строк "
                f"({itog_sum}) не равна строке ИТОГО "
                f"({report['ИТОГО'][1]})",
            )
        ws[column[5] + self.row] = itog_sum
        ws[column[5] + self.row].number_format = "#,##0.00 ₽"
        merge_table_bold()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if report["Дата"][0] == report["Дата"][1] - timedelta(1):
            date_ = datetime.strftime(report["Дата"][0], "%Y-%m-%d")
        else:
            date_ = (
                f"{datetime.strftime(report['Дата'][0], '%Y-%m-%d')} - "
                f"{datetime.strftime(report['Дата'][1], '%Y-%m-%d')}"
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f" Отчет платежного агента {report['Организация'][1]}"
            + ".xlsx"
        )
        logger.info(f"Сохранение отчета платежного агента {report['Организация'][1]} в {path}")
        path = self.create_path(path, date_from)
        self.save_file(path, wb)
        return path

    def save_organisation_total(self, itog_report, date_from):
        """
        Сохраняет Итоговый отчет в Excel
        """
        organisation_total = {}
        for key in itog_report:
            organisation_total[itog_report[key][3]] = {}
        for key in itog_report:
            organisation_total[itog_report[key][3]][itog_report[key][2]] = []
        for key in itog_report:
            organisation_total[itog_report[key][3]][itog_report[key][2]].append(
                (key, itog_report[key][0], itog_report[key][1])
            )

        column = ["", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Итоговый отчет"
        # шрифты
        ws["C1"].font = ReportStyle.h1
        # выравнивание
        ws["C1"].alignment = ReportStyle.align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 8
        ws.column_dimensions["B"].width = 1 / 7 * 8
        ws.column_dimensions["C"].width = 1 / 7 * 80
        ws.column_dimensions["D"].width = 1 / 7 * 8
        ws.column_dimensions["E"].width = 1 / 7 * 88
        ws.column_dimensions["F"].width = 1 / 7 * 8
        ws.column_dimensions["G"].width = 1 / 7 * 24
        ws.column_dimensions["H"].width = 1 / 7 * 8
        ws.column_dimensions["I"].width = 1 / 7 * 80
        ws.column_dimensions["J"].width = 1 / 7 * 8
        ws.column_dimensions["K"].width = 1 / 7 * 144
        ws.column_dimensions["L"].width = 1 / 7 * 144
        ws.column_dimensions["M"].width = 1 / 7 * 8

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[3] + next_row()] = "Итоговый отчет"
        ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=12)
        ws[column[1] + next_row()] = ""
        ws[column[3] + next_row()] = organisation_total["Организация"]["Организация"][0][0]
        ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=12)
        ws[column[3] + self.row].font = ReportStyle.font
        ws[column[3] + self.row].alignment = ReportStyle.align_top
        ws[column[1] + next_row()] = ""

        ws[column[3] + next_row()] = "За период с:"
        ws[column[3] + self.row].font = ReportStyle.font
        ws[column[3] + self.row].alignment = ReportStyle.align_top
        ws[column[5] + self.row] = itog_report["Дата"][0].strftime("%d.%m.%Y")
        ws[column[5] + self.row].font = ReportStyle.font_bold
        ws[column[5] + self.row].alignment = ReportStyle.align_top
        ws[column[7] + self.row] = "По:"
        ws[column[7] + self.row].font = ReportStyle.font
        ws[column[7] + self.row].alignment = ReportStyle.align_top
        ws[column[9] + self.row] = (itog_report["Дата"][1] - timedelta(1)).strftime("%d.%m.%Y")
        ws[column[9] + self.row].font = ReportStyle.font_bold
        ws[column[9] + self.row].alignment = ReportStyle.align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.font
            ws[column[10] + self.row].font = ReportStyle.font
            ws[column[12] + self.row].font = ReportStyle.font
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[12] + self.row].alignment = ReportStyle.align_top
            b = 2
            while b <= 13:
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        def merge_table_h3():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.h3
            ws[column[10] + self.row].font = ReportStyle.h3
            ws[column[12] + self.row].font = ReportStyle.h3
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[12] + self.row].alignment = ReportStyle.align_top
            ws[column[2] + self.row].border = ReportStyle.border_left
            ws[column[13] + self.row].border = ReportStyle.border_right

        def merge_table_h2():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.h2
            ws[column[10] + self.row].font = ReportStyle.h2
            ws[column[12] + self.row].font = ReportStyle.h2
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[12] + self.row].alignment = ReportStyle.align_top
            ws[column[2] + self.row].border = ReportStyle.border_left
            ws[column[13] + self.row].border = ReportStyle.border_right

        def merge_width_h2():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.h2
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            b = 2
            while b <= 13:
                if b == 2:
                    ws[column[b] + self.row].border = ReportStyle.border_left_top
                elif b == 13:
                    ws[column[b] + self.row].border = ReportStyle.border_right_top
                else:
                    ws[column[b] + self.row].border = ReportStyle.border_top
                b += 1

        def merge_width_h3():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = ReportStyle.h3
            ws[column[2] + self.row].alignment = ReportStyle.align_top
            b = 2
            while b <= 13:
                if b == 2:
                    ws[column[b] + self.row].border = ReportStyle.border_left
                elif b == 13:
                    ws[column[b] + self.row].border = ReportStyle.border_right
                b += 1

        ws[column[2] + next_row()] = "Название"
        ws[column[10] + self.row] = "Количество"
        ws[column[12] + self.row] = "Сумма"
        merge_table()
        ws[column[2] + self.row].font = ReportStyle.h3
        ws[column[10] + self.row].font = ReportStyle.h3
        ws[column[12] + self.row].font = ReportStyle.h3
        ws[column[2] + self.row].alignment = ReportStyle.align_top
        ws[column[10] + self.row].alignment = ReportStyle.align_top
        ws[column[12] + self.row].alignment = ReportStyle.align_top

        groups = [
            "Депозит",
            "Карты",
            "Услуги",
            "Товары",
            "Платные зоны",
        ]
        all_count = 0
        all_sum = 0
        try:
            for gr in groups:
                ws[column[2] + next_row()] = gr
                merge_width_h2()
                group_count = 0
                group_sum = 0
                organisation_total_groups = organisation_total.get(gr)
                if not organisation_total_groups:
                    continue
                for group in organisation_total_groups:
                    ws[column[2] + next_row()] = group
                    merge_width_h3()
                    service_count = 0
                    service_sum = 0
                    servises = organisation_total_groups.get(group, [])
                    if not servises:
                        continue
                    for service in servises:
                        try:
                            service_count += service[1]
                            service_sum += service[2]
                        except TypeError:
                            pass
                        ws[column[2] + next_row()] = service[0]
                        ws[column[10] + self.row] = service[1]
                        ws[column[12] + self.row] = service[2]
                        ws[column[12] + self.row].number_format = "#,##0.00 ₽"
                        merge_table()
                    ws[column[10] + next_row()] = service_count
                    ws[column[12] + self.row] = service_sum
                    ws[column[12] + self.row].number_format = "#,##0.00 ₽"
                    merge_table_h3()
                    group_count += service_count
                    group_sum += service_sum
                ws[column[10] + next_row()] = group_count
                ws[column[12] + self.row] = group_sum
                ws[column[12] + self.row].number_format = "#,##0.00 ₽"
                merge_table_h2()
                all_count += group_count
                all_sum += group_sum
                group_count = 0
                group_sum = 0
        except KeyError:
            pass

        bars_total_sum = organisation_total["Итого по отчету"][""][0]
        if all_sum == bars_total_sum[2]:
            ws[column[2] + next_row()] = bars_total_sum[0]
            ws[column[10] + self.row] = bars_total_sum[1]
            ws[column[12] + self.row] = bars_total_sum[2]
            self.total_report_sum = all_sum
        else:
            error_code = "Ошибка: Итоговые суммы не совпадают."
            error_message = (
                f'"Итого по отчету" из Барса ({bars_total_sum[2]})'
                f" не совпадает с итоговой суммой по формируемым строкам ({all_sum})."
            )
            logger.error(f"{error_code} {error_message}")
            return None

        ws[column[12] + self.row].number_format = "#,##0.00 ₽"
        merge_table_h2()
        ws[column[2] + self.row].alignment = ReportStyle.align_bottom
        ws[column[10] + self.row].alignment = ReportStyle.align_bottom
        ws[column[12] + self.row].alignment = ReportStyle.align_bottom
        b = 2
        while b <= 13:
            ws[column[b] + self.row].border = ReportStyle.border_top_bottom
            b += 1
        end_line = int(self.row)

        # раскрвшивание фона для заголовков
        i = 2
        while i <= 13:
            ws[column[i] + "6"].fill = ReportStyle.fill
            i += 1

        # обводка
        # ws['A3'].border = ReportStyle.border

        # вручную устанавливаем высоту первой строки
        rd = ws.row_dimensions[1]
        rd.height = 21.75

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1

        # Высота строк
        ws.row_dimensions[2].height = 5.25
        ws.row_dimensions[4].height = 6.75
        ws.row_dimensions[end_line].height = 30.75

        # выравнивание столбца
        for cellObj in ws["A2:A5"]:
            for cell in cellObj:
                ws[cell.coordinate].alignment = ReportStyle.align_left

        # перетягивание ячеек
        # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
        # dims = {}
        # for row in ws.rows:
        #     for cell in row:
        #         if cell.value:
        #             dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        # for col, value in dims.items():
        #     # value * коэфициент
        #     ws.column_dimensions[col].width = value * 1.5

        # сохранение файла в текущую директорию
        if itog_report["Дата"][0] == itog_report["Дата"][1] - timedelta(1):
            date_ = datetime.strftime(itog_report["Дата"][0], "%Y-%m-%d")
        else:
            date_ = (
                f"{datetime.strftime(itog_report['Дата'][0], '%Y-%m-%d')} - "
                f"{datetime.strftime(itog_report['Дата'][1] - timedelta(1), '%Y-%m-%d')}"
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f" Итоговый отчет по {organisation_total['Организация']['Организация'][0][0]} "
            + ".xlsx"
        )
        logger.info(f"Сохранение Итогового отчета по {organisation_total['Организация']['Организация'][0][0]} в {path}")
        path = self.create_path(path, date_from)
        self.save_file(path, wb)
        return path

    def save_cashdesk_report(self, cashdesk_report, date_from):
        """Сохраняет Суммовой отчет в Excel"""

        column = [
            "",
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
        ]

        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Суммовой отчет по чековой ленте"
        # шрифты
        ws["A1"].font = ReportStyle.h1
        # выравнивание
        ws["A1"].alignment = ReportStyle.align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 124
        ws.column_dimensions["B"].width = 1 / 7 * 88
        ws.column_dimensions["C"].width = 1 / 7 * 28
        ws.column_dimensions["D"].width = 1 / 7 * 24
        ws.column_dimensions["E"].width = 1 / 7 * 32
        ws.column_dimensions["F"].width = 1 / 7 * 1
        ws.column_dimensions["G"].width = 1 / 7 * 79
        ws.column_dimensions["H"].width = 1 / 7 * 3
        ws.column_dimensions["I"].width = 1 / 7 * 5
        ws.column_dimensions["J"].width = 1 / 7 * 96
        ws.column_dimensions["K"].width = 1 / 7 * 88
        ws.column_dimensions["L"].width = 1 / 7 * 8
        ws.column_dimensions["M"].width = 1 / 7 * 80
        ws.column_dimensions["N"].width = 1 / 7 * 96

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = "Суммовой отчет по чековой ленте"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        # шрифты
        ws[column[1] + self.row].font = ReportStyle.h1
        # выравнивание
        ws[column[1] + self.row].alignment = ReportStyle.align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f"{cashdesk_report['Организация'][0][0]}"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        ws[column[1] + self.row].font = ReportStyle.font
        ws[column[1] + self.row].alignment = ReportStyle.align_top

        ws[column[1] + next_row()] = "За период с:"
        ws[column[1] + self.row].font = ReportStyle.font
        ws[column[1] + self.row].alignment = ReportStyle.align_top
        ws[column[2] + self.row] = cashdesk_report["Дата"][0][0].strftime("%d.%m.%Y")
        ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=3)
        ws[column[2] + self.row].font = ReportStyle.font_bold
        ws[column[2] + self.row].alignment = ReportStyle.align_top
        ws[column[4] + self.row] = "по"
        ws[column[4] + self.row].font = ReportStyle.font
        ws[column[4] + self.row].alignment = ReportStyle.align_top
        ws[column[5] + self.row] = (cashdesk_report["Дата"][0][1] - timedelta(1)).strftime("%d.%m.%Y")
        ws.merge_cells(start_row=self.row, start_column=5, end_row=self.row, end_column=7)
        ws[column[5] + self.row].font = ReportStyle.font_bold
        ws[column[5] + self.row].alignment = ReportStyle.align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=6)
            ws.merge_cells(start_row=self.row, start_column=7, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=11, end_row=self.row, end_column=12)
            ws[column[1] + self.row].font = ReportStyle.font
            ws[column[3] + self.row].font = ReportStyle.font
            ws[column[7] + self.row].font = ReportStyle.font
            ws[column[10] + self.row].font = ReportStyle.font
            ws[column[11] + self.row].font = ReportStyle.font
            ws[column[13] + self.row].font = ReportStyle.font
            ws[column[14] + self.row].font = ReportStyle.font
            ws[column[1] + self.row].alignment = ReportStyle.align_top
            ws[column[3] + self.row].alignment = ReportStyle.align_top
            ws[column[7] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[11] + self.row].alignment = ReportStyle.align_top
            ws[column[13] + self.row].alignment = ReportStyle.align_top
            ws[column[14] + self.row].alignment = ReportStyle.align_top
            ws[column[3] + self.row].number_format = "#,##0.00 ₽"
            ws[column[7] + self.row].number_format = "#,##0.00 ₽"
            ws[column[10] + self.row].number_format = "#,##0.00 ₽"
            ws[column[11] + self.row].number_format = "#,##0.00 ₽"
            ws[column[13] + self.row].number_format = "#,##0.00 ₽"
            ws[column[14] + self.row].number_format = "#,##0.00 ₽"
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        def merge_table_h3():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=6)
            ws.merge_cells(start_row=self.row, start_column=7, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=11, end_row=self.row, end_column=12)
            ws[column[1] + self.row].font = ReportStyle.h3
            ws[column[3] + self.row].font = ReportStyle.h3
            ws[column[7] + self.row].font = ReportStyle.h3
            ws[column[10] + self.row].font = ReportStyle.h3
            ws[column[11] + self.row].font = ReportStyle.h3
            ws[column[13] + self.row].font = ReportStyle.h3
            ws[column[14] + self.row].font = ReportStyle.h3
            ws[column[1] + self.row].alignment = ReportStyle.align_top
            ws[column[3] + self.row].alignment = ReportStyle.align_top
            ws[column[7] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[11] + self.row].alignment = ReportStyle.align_top
            ws[column[13] + self.row].alignment = ReportStyle.align_top
            ws[column[14] + self.row].alignment = ReportStyle.align_top
            ws[column[3] + self.row].number_format = "#,##0.00 ₽"
            ws[column[7] + self.row].number_format = "#,##0.00 ₽"
            ws[column[10] + self.row].number_format = "#,##0.00 ₽"
            ws[column[11] + self.row].number_format = "#,##0.00 ₽"
            ws[column[13] + self.row].number_format = "#,##0.00 ₽"
            ws[column[14] + self.row].number_format = "#,##0.00 ₽"
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        def merge_table_red():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=6)
            ws.merge_cells(start_row=self.row, start_column=7, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=11, end_row=self.row, end_column=12)
            ws[column[1] + self.row].font = ReportStyle.font_red
            ws[column[3] + self.row].font = ReportStyle.font_red
            ws[column[7] + self.row].font = ReportStyle.font_red
            ws[column[10] + self.row].font = ReportStyle.font_red
            ws[column[11] + self.row].font = ReportStyle.font_red
            ws[column[13] + self.row].font = ReportStyle.font_red
            ws[column[14] + self.row].font = ReportStyle.font_red
            ws[column[1] + self.row].alignment = ReportStyle.align_top
            ws[column[3] + self.row].alignment = ReportStyle.align_top
            ws[column[7] + self.row].alignment = ReportStyle.align_top
            ws[column[10] + self.row].alignment = ReportStyle.align_top
            ws[column[11] + self.row].alignment = ReportStyle.align_top
            ws[column[13] + self.row].alignment = ReportStyle.align_top
            ws[column[14] + self.row].alignment = ReportStyle.align_top
            ws[column[3] + self.row].number_format = "#,##0.00 ₽"
            ws[column[7] + self.row].number_format = "#,##0.00 ₽"
            ws[column[10] + self.row].number_format = "#,##0.00 ₽"
            ws[column[11] + self.row].number_format = "#,##0.00 ₽"
            ws[column[13] + self.row].number_format = "#,##0.00 ₽"
            ws[column[14] + self.row].number_format = "#,##0.00 ₽"
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        def merge_width_red():
            ws.merge_cells(
                start_row=self.row,
                start_column=1,
                end_row=self.row,
                end_column=len(column) - 1,
            )
            ws[column[1] + self.row].font = ReportStyle.font_red
            ws[column[1] + self.row].alignment = ReportStyle.align_top
            b = 1
            while b < len(column):
                if b == 1:
                    ws[column[b] + self.row].border = ReportStyle.border_left
                elif b == len(column) - 1:
                    ws[column[b] + self.row].border = ReportStyle.border_right
                else:
                    ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        ws[column[1] + next_row()] = "Касса №"
        ws[column[3] + self.row] = "Сумма"
        ws[column[7] + self.row] = "Наличными"
        ws[column[10] + self.row] = "Безналичными"
        ws[column[11] + self.row] = "Со счета"
        ws[column[13] + self.row] = "Бонусами"
        ws[column[14] + self.row] = "MaxBonux"
        merge_table_h3()
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = ReportStyle.fill
            b += 1

        for typpe in cashdesk_report:
            if typpe not in {"Дата", "Организация"}:
                if typpe != "Итого":
                    ws[column[1] + next_row()] = typpe
                    merge_width_red()
                for line in cashdesk_report[typpe]:
                    ws[column[1] + next_row()] = line[0]
                    ws[column[3] + self.row] = line[1]
                    ws[column[7] + self.row] = line[2]
                    ws[column[10] + self.row] = line[3]
                    ws[column[11] + self.row] = line[4]
                    ws[column[13] + self.row] = line[5]
                    ws[column[14] + self.row] = line[6] - line[3]
                    if line[0] == "Итого":
                        merge_table_red()
                    elif line[0] == "Итого по отчету":
                        merge_table_h3()
                    else:
                        merge_table()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if cashdesk_report["Дата"][0][0] == cashdesk_report["Дата"][0][1] - timedelta(1):
            date_ = datetime.strftime(cashdesk_report["Дата"][0][0], "%Y-%m-%d")
        else:
            date_ = (
                f"{datetime.strftime(cashdesk_report['Дата'][0][0], '%Y-%m-%d')} - "
                f"{datetime.strftime(cashdesk_report['Дата'][0][1] - timedelta(1), '%Y-%m-%d')}"
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f" Суммовой отчет по {cashdesk_report['Организация'][0][0]}"
            + ".xlsx"
        )
        logger.info(f"Сохранение Суммового отчета по {cashdesk_report['Организация'][0][0]} в {path}")
        path = self.create_path(path, date_from)
        self.save_file(path, wb)
        return path

    def save_client_count_totals(self, client_count_totals_org, date_from):
        """
        Сохраняет отчет по количеству клиентов за день в Excel
        """

        column = ["", "A", "B", "C", "D", "E"]

        self.row = "0"

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = "Количество человек за день"
        # шрифты
        ws["A1"].font = ReportStyle.h1
        # выравнивание
        ws["A1"].alignment = ReportStyle.align_left

        # Ширина стролбцов
        ws.column_dimensions["A"].width = 1 / 7 * 124
        ws.column_dimensions["B"].width = 1 / 7 * 21
        ws.column_dimensions["C"].width = 1 / 7 * 95
        ws.column_dimensions["D"].width = 1 / 7 * 24
        ws.column_dimensions["E"].width = 1 / 7 * 80

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = "Количество человек за день"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        # шрифты
        ws[column[1] + self.row].font = ReportStyle.h1
        # выравнивание
        ws[column[1] + self.row].alignment = ReportStyle.align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f"{client_count_totals_org[0][0]}"
        ws.merge_cells(
            start_row=self.row,
            start_column=1,
            end_row=self.row,
            end_column=len(column) - 1,
        )
        ws[column[1] + self.row].font = ReportStyle.font
        ws[column[1] + self.row].alignment = ReportStyle.align_top

        ws[column[1] + next_row()] = "За период с:"
        ws[column[1] + self.row].font = ReportStyle.font
        ws[column[1] + self.row].alignment = ReportStyle.align_top
        ws[column[2] + self.row] = client_count_totals_org[1][0].strftime("%d.%m.%Y")
        ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=3)
        ws[column[2] + self.row].font = ReportStyle.font_bold
        ws[column[2] + self.row].alignment = ReportStyle.align_top
        ws[column[4] + self.row] = "по"
        ws[column[4] + self.row].font = ReportStyle.font
        ws[column[4] + self.row].alignment = ReportStyle.align_top
        ws[column[5] + self.row] = (client_count_totals_org[-2][0]).strftime("%d.%m.%Y")
        ws.merge_cells(start_row=self.row, start_column=5, end_row=self.row, end_column=7)
        ws[column[5] + self.row].font = ReportStyle.font_bold
        ws[column[5] + self.row].alignment = ReportStyle.align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=5)
            ws[column[1] + self.row].font = ReportStyle.font
            ws[column[3] + self.row].font = ReportStyle.font
            ws[column[1] + self.row].alignment = ReportStyle.align_top
            ws[column[3] + self.row].alignment = ReportStyle.align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        def merge_table_bold():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=5)
            ws[column[1] + self.row].font = ReportStyle.font_bold
            ws[column[3] + self.row].font = ReportStyle.font_bold
            ws[column[1] + self.row].alignment = ReportStyle.align_top
            ws[column[3] + self.row].alignment = ReportStyle.align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = ReportStyle.border
                b += 1

        ws[column[1] + next_row()] = "Дата"
        ws[column[3] + self.row] = "Количество клиентов"
        merge_table_bold()
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = ReportStyle.fill
            b += 1

        for line in client_count_totals_org:
            try:
                ws[column[1] + next_row()] = line[0].strftime("%d.%m.%Y")
                ws[column[3] + self.row] = line[1]
                merge_table()
            except AttributeError:
                pass

        ws[column[1] + next_row()] = "Итого"
        ws[column[3] + self.row] = client_count_totals_org[-1][1]
        merge_table_bold()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if client_count_totals_org[0][1]:
            date_ = datetime.strftime(client_count_totals_org[1][0], "%Y-%m")
        else:
            date_ = (
                f"{datetime.strftime(client_count_totals_org[1][0], '%Y-%m-%d')} - "
                f"{datetime.strftime(client_count_totals_org[-2][0], '%Y-%m-%d')}"
            )
        path = (
            settings.local_folder
            + settings.report_path
            + date_
            + f" Количество клиентов за день по {client_count_totals_org[0][0]}"
            + ".xlsx"
        )
        logger.info(f"Сохранение отчета по количеству клиентов по {client_count_totals_org[0][0]} в {path}")
        path = self.create_path(path, date_from)
        self.save_file(path, wb)
        return path

    def create_path(self, path: str, date_from: datetime):
        """Проверяет наличие указанного пути. В случае отсутствия каких-либо папок создает их."""

        logger.info("Проверка локальных путей сохранения файлов...")
        list_path = path.split("/")
        path = ""
        end_path = ""
        if list_path[-1][-4:] == ".xls" or list_path[-1]:
            end_path = list_path.pop()
        list_path.append(date_from.strftime("%Y"))
        list_path.append(date_from.strftime("%m") + "-" + date_from.strftime("%B"))
        directory = os.getcwd()
        for folder in list_path:
            if folder not in os.listdir():
                os.mkdir(folder)
                logger.debug(f'В директории "{os.getcwd()}" создана папка "{folder}"')
                os.chdir(folder)
            else:
                os.chdir(folder)
            path += folder + "/"
        path += end_path
        os.chdir(directory)
        return path

    def save_file(self, path, file):
        """Проверяет не занят ли файл другим процессом и если нет, то перезаписывает его, в противном
        случае выводит диалоговое окно с предложением закрыть файл и продолжить.
        """
        try:
            file.save(path)
        except PermissionError as e:
            logger.error(f'Файл "{path}" занят другим процессом.\n{e!r}')

    def create_path_yadisk(self, path, date_from):
        """Проверяет наличие указанного пути в Яндекс Диске.

        В случае отсутствия каких-либо папок создает их.
        """
        logger.info("Проверка путей сохранения файлов на Яндекс.Диске...")
        list_path = path.split("/")
        if list_path[-1][-4:] == ".xls" or list_path[-1] == "":
            list_path.pop()
        list_path.append(date_from.strftime("%Y"))
        list_path.append(date_from.strftime("%m") + "-" + date_from.strftime("%B"))
        directory = "/"
        list_path_yandex = []
        for folder in list_path:
            folder = directory + folder
            directory = folder + "/"
            list_path_yandex.append(folder)
        directory = "/"
        for folder in list_path_yandex:
            folders_list = []
            folders_list_yandex = list(self._yadisk.listdir(directory))
            for key in folders_list_yandex:
                if not key["file"]:
                    folders_list.append(directory + key["name"])

            if folder not in folders_list:
                self._yadisk.mkdir(folder)
                logger.info(f'Создание новой папки в YandexDisk - "{folder}"')
                directory = folder + "/"
            else:
                directory = folder + "/"
        return list_path_yandex[-1] + "/"

    def sync_to_yadisk(self, paths: list[str], date_from: datetime) -> list[SyncResourceLinkObject]:
        """Копирует локальные файлы в Яндекс Диск."""

        logger.info("Копирование отчетов в Яндекс.Диск...")
        if not paths:
            logger.warning("Нет ни одного отчета для отправки в Yandex.Disk")
            return []

        links = []
        if self._yadisk.check_token():
            yadisk_path = "" + settings.report_path
            remote_folder = self.create_path_yadisk(yadisk_path, date_from)
            for path in paths:
                if path is None:
                    continue

                remote_path = remote_folder + path.split("/")[-1]
                file_name = f"'{path.split('/')[-1]}'"
                files_list_yandex = list(self._yadisk.listdir(remote_folder))
                files_list = []
                for key in files_list_yandex:
                    if key["file"]:
                        files_list.append(remote_folder + key["name"])  # PERF401

                if remote_path in files_list:
                    logger.warning(f"Файл {file_name} уже существует в '{remote_folder}' и будет заменен!")
                    self._yadisk.remove(remote_path, permanently=True)

                links.append(self._yadisk.upload(path, remote_path))

            return links

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ошибка YaDisk: token не валиден",
        )

    async def read_yadisk_files(self, root_path: str, date_from: datetime, file_template_name: str):
        yield self._async_yadisk.listdir(root_path)


def get_yandex_repo() -> YandexRepository:
    return YandexRepository(token=settings.yadisk_token)
